from __future__ import annotations

import hashlib
import json
import logging
import re
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, timedelta, timezone
from typing import TYPE_CHECKING, Any, TypedDict

logger = logging.getLogger(__name__)

from sqlalchemy import func
from sqlalchemy import true as sa_true
from sqlalchemy.orm import Session
from sqlalchemy.orm.query import RowReturningQuery

if TYPE_CHECKING:
    from app.models.contact import Contact
    from app.models.member import Member
    from app.models.p4x_fee import P4xFee
    from app.models.p4x_specialcontact import P4xSpecialcontact

from app.models.p4x_account import P4xAccount
from app.models.p4x_category import P4xCategory
from app.models.p4x_category_direct import P4xCategoryDirect
from app.models.p4x_category_filter import P4xCategoryFilter
from app.models.p4x_category_filter_hit import P4xCategoryFilterHit
from app.models.p4x_partner import P4xPartner
from app.models.p4x_transaction import P4xTransaction

FEE_CATEGORY_ID = 1
SUMUP_ACCOUNT_ID = 1
SUMUP_CATEGORY_NAME = "projekt.bude.sumup"
GEORGE_BIC = "GIBAATWWXXX"
INACTIVE_THRESHOLD_DAYS = 730
PAGINATION_SIZE = 100


class FeeBalanceResult(TypedDict):
    """Typed result of calculate_fee_balance."""

    start_date: str
    start_balance: float
    count: dict[str, int]
    sum: dict[str, float]
    end_date: str
    end_balance: float
    progress: list[dict[str, str | float]]


# ---------------------------------------------------------------------------
# Parser
# ---------------------------------------------------------------------------


@dataclass
class ParseResult:
    success: bool
    message: str
    entries: list[dict[str, Any]] = field(default_factory=list)


def parse_george_json(bic: str, raw_json: str) -> ParseResult:  # noqa: C901
    if bic != GEORGE_BIC:
        return ParseResult(
            success=False, message=f"No parser method found for BIC {bic}"
        )

    try:
        data = json.loads(raw_json)
    except (json.JSONDecodeError, TypeError):
        return ParseResult(success=False, message="failed to parse given raw json data")

    if not isinstance(data, list):
        return ParseResult(success=False, message="failed to parse given raw json data")

    result: list[dict[str, Any]] = []

    for struct in data:
        if "booking" not in struct:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: booking"
                ),
            )
        if "valuation" not in struct:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: valuation"
                ),
            )
        if "partnerAccount" not in struct:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: partnerAccount"
                ),
            )
        if not isinstance(struct["partnerAccount"], dict):
            struct["partnerAccount"] = {"iban": ""}
        if "iban" not in struct["partnerAccount"]:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: partnerAccount.iban"
                ),
            )
        if "amount" not in struct:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: amount"
                ),
            )
        if "value" not in struct["amount"]:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: amount.value"
                ),
            )
        if "precision" not in struct["amount"]:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: amount.precision"
                ),
            )
        if "reference" not in struct and "receiverReference" not in struct:
            return ParseResult(
                success=False,
                message=(
                    "given data contains at least one corrupt entry:"
                    " missing field: reference or receiverReference"
                ),
            )

        ref = struct.get("reference") or ""
        recv = struct.get("receiverReference") or ""
        if ref and recv:
            subject = ref if len(ref) > len(recv) else recv
        elif ref:
            subject = ref
        elif recv:
            subject = recv
        else:
            subject = ""

        precision = int(str(struct["amount"]["precision"]).strip())
        raw_value = float(str(struct["amount"]["value"]).strip())
        amount = raw_value / (10**precision)

        result.append(
            {
                "payload": {
                    "booking": _parse_date_string(str(struct["booking"]).strip()),
                    "valuation": _parse_date_string(str(struct["valuation"]).strip()),
                    "iban": str(struct["partnerAccount"]["iban"]).strip(),
                    "amount": f"{amount:.2f}",
                    "subject": subject.strip(),
                },
                "raw": json.dumps(struct),
            }
        )

    return ParseResult(success=True, message="finished successfully", entries=result)


def _parse_date_string(date_str: str) -> date:
    return date.fromisoformat(date_str[:10])


# ---------------------------------------------------------------------------
# SHA256 hash (must match PHP Carbon toJSON + json_encode behavior)
# ---------------------------------------------------------------------------


def _date_to_carbon_json(booking_date: date, original_date_str: str) -> str:
    """Convert a date to Carbon 3's toJSON() UTC format.

    Carbon::parse() preserves the original timezone, then toJSON() converts to
    UTC and formats as 'YYYY-MM-DDTHH:MM:SS.000000Z'.
    """
    m = re.match(
        r"(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})(?:\.\d+)?([+-])(\d{2})(\d{2})",
        original_date_str.strip(),
    )
    if not m:
        return f"{booking_date.isoformat()}T00:00:00.000000Z"

    dt = datetime.fromisoformat(f"{m.group(1)}T{m.group(2)}")
    sign = 1 if m.group(3) == "+" else -1
    tz = timezone(
        timedelta(hours=sign * int(m.group(4)), minutes=sign * int(m.group(5)))
    )
    dt_utc = dt.replace(tzinfo=tz).astimezone(UTC)
    return dt_utc.strftime("%Y-%m-%dT%H:%M:%S") + ".000000Z"


def _php_json_encode(obj: list[str]) -> str:
    """Replicate PHP's default json_encode behavior."""
    s = json.dumps(obj, ensure_ascii=True, separators=(",", ":"))
    return s.replace("/", "\\/")


def compute_transaction_hash(
    booking_carbon_json: str,
    valuation_carbon_json: str,
    iban: str,
    amount_str: str,
    subject: str,
) -> str:
    payload = _php_json_encode(
        [
            booking_carbon_json,
            valuation_carbon_json,
            iban,
            amount_str,
            subject,
        ]
    )
    return hashlib.sha256(payload.encode()).hexdigest()


# ---------------------------------------------------------------------------
# Import
# ---------------------------------------------------------------------------


def import_transactions(
    db: Session,
    account: P4xAccount,
    parsed_entries: list[dict[str, Any]],
    original_structs: list[dict[str, Any]],
) -> dict[str, int]:
    """Import parsed transactions into the database.

    Returns a summary dict with counts by status.
    """
    summary: dict[str, int] = {}

    for i, entry in enumerate(parsed_entries):
        payload = entry["payload"]
        raw = entry["raw"]

        summary["giventotal"] = summary.get("giventotal", 0) + 1

        mandatory = {"booking", "valuation", "iban", "amount", "subject"}
        if not mandatory.issubset(payload.keys()):
            summary["error"] = summary.get("error", 0) + 1
            continue

        amount_float = float(payload["amount"])
        if round(amount_float, 3) == 0.0:
            summary["zero_skipped"] = summary.get("zero_skipped", 0) + 1
            continue

        orig_struct = original_structs[i]
        booking_carbon = _date_to_carbon_json(
            payload["booking"], str(orig_struct.get("booking", ""))
        )
        valuation_carbon = _date_to_carbon_json(
            payload["valuation"], str(orig_struct.get("valuation", ""))
        )

        sha256hash = compute_transaction_hash(
            booking_carbon,
            valuation_carbon,
            payload["iban"],
            payload["amount"],
            payload["subject"],
        )

        if payload["booking"] < account.init_date:
            db.query(P4xTransaction).filter(
                P4xTransaction.sha256hash == sha256hash,
            ).update({"deleted_at": datetime.now(UTC)})
            db.commit()
            summary["before_init_date"] = summary.get("before_init_date", 0) + 1
            continue

        existing = (
            db.query(P4xTransaction)
            .filter(
                P4xTransaction.sha256hash == sha256hash,
                P4xTransaction.deleted_at.is_(None),
            )
            .first()
        )

        if existing:
            status = "existing"
            existing.booking = payload["booking"]
            existing.valuation = payload["valuation"]
            existing.iban = payload["iban"]
            existing.amount = amount_float
            existing.subject = payload["subject"]
            existing.raw = raw
            if existing.p4x_account_id != account.id:
                existing.p4x_account_id = account.id
                status = "existing_with_new_binding"
            existing.updated_at = datetime.now(UTC)
        else:
            status = "new"
            tx = P4xTransaction(
                sha256hash=sha256hash,
                booking=payload["booking"],
                valuation=payload["valuation"],
                iban=payload["iban"],
                amount=amount_float,
                subject=payload["subject"],
                p4x_account_id=account.id,
                raw=raw,
                created_at=datetime.now(UTC),
                updated_at=datetime.now(UTC),
            )
            db.add(tx)

        summary[status] = summary.get(status, 0) + 1

    db.commit()
    return summary


# ---------------------------------------------------------------------------
# Account queries
# ---------------------------------------------------------------------------


def get_account_balance(
    db: Session,
    account: P4xAccount,
    up_to_date: date | None = None,
) -> float:
    if up_to_date is None:
        up_to_date = datetime.now(UTC).date()

    total = (
        db.query(func.sum(P4xTransaction.amount))
        .filter(
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.booking <= up_to_date,
            P4xTransaction.deleted_at.is_(None),
        )
        .scalar()
    ) or 0.0

    return round(float(account.init_balance) + total, 2)


def get_transactions_by_month(
    db: Session,
    account: P4xAccount,
    year: int,
    month: int,
    page: int,
) -> tuple[list[P4xTransaction], int]:
    query = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
            func.substr(P4xTransaction.booking, 1, 4) == str(year),
            func.substr(P4xTransaction.booking, 6, 2) == f"{month:02d}",
        )
        .order_by(P4xTransaction.booking.desc())
    )
    total = query.count()
    items = query.offset((page - 1) * PAGINATION_SIZE).limit(PAGINATION_SIZE).all()
    return items, total


def get_transactions_by_partner(
    db: Session,
    account: P4xAccount,
    partner_type: str,
    partner_id: int,
    page: int,
) -> tuple[list[P4xTransaction], int]:
    partner_ibans = [
        r[0]
        for r in db.query(P4xPartner.iban)
        .filter(
            P4xPartner.partner_type == partner_type,
            P4xPartner.partner_id == partner_id,
            P4xPartner.deleted_at.is_(None),
        )
        .all()
    ]

    query = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
        )
        .filter(
            # (partner matches AND no delegating) OR (delegating matches)
            (
                P4xTransaction.iban.in_(partner_ibans)
                & P4xTransaction.delegating_partner_type.is_(None)
            )
            | (
                (P4xTransaction.delegating_partner_type == partner_type)
                & (P4xTransaction.delegating_partner_id == partner_id)
            )
        )
        .order_by(P4xTransaction.booking.desc())
    )
    total = query.count()
    items = query.offset((page - 1) * PAGINATION_SIZE).limit(PAGINATION_SIZE).all()
    return items, total


def get_transactions_by_category(
    db: Session,
    account: P4xAccount,
    category_id: int,
    page: int,
) -> tuple[list[P4xTransaction], int]:
    direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.p4x_category_id == category_id,
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }

    filter_ids = [
        r[0]
        for r in db.query(P4xCategoryFilter.id)
        .filter(
            P4xCategoryFilter.p4x_category_id == category_id,
        )
        .all()
    ]
    filter_tx_ids = (
        {
            r[0]
            for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
            .filter(
                P4xCategoryFilterHit.p4x_category_filter_id.in_(filter_ids),
            )
            .all()
        }
        if filter_ids
        else set()
    )

    all_direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }
    filter_only_tx_ids = filter_tx_ids - all_direct_tx_ids
    all_tx_ids = direct_tx_ids | filter_only_tx_ids

    if not all_tx_ids:
        return [], 0

    query = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
            P4xTransaction.id.in_(all_tx_ids),
        )
        .order_by(P4xTransaction.booking.desc())
    )
    total = query.count()
    items = query.offset((page - 1) * PAGINATION_SIZE).limit(PAGINATION_SIZE).all()
    return items, total


def get_transactions_by_filter(
    db: Session,
    account: P4xAccount,
    filter_id: int,
    page: int,
) -> tuple[list[P4xTransaction], int]:
    hit_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
        .filter(
            P4xCategoryFilterHit.p4x_category_filter_id == filter_id,
        )
        .all()
    }

    all_direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }
    tx_ids = hit_tx_ids - all_direct_tx_ids

    if not tx_ids:
        return [], 0

    query = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
            P4xTransaction.id.in_(tx_ids),
        )
        .order_by(P4xTransaction.booking.desc())
    )
    total = query.count()
    items = query.offset((page - 1) * PAGINATION_SIZE).limit(PAGINATION_SIZE).all()
    return items, total


# ---------------------------------------------------------------------------
# Warnings
# ---------------------------------------------------------------------------


def get_warnings_partner(
    db: Session,
    limit: int | None = None,
) -> tuple[list[P4xTransaction], int]:
    partner_ibans = {
        r[0]
        for r in db.query(P4xPartner.iban)
        .filter(
            P4xPartner.deleted_at.is_(None),
        )
        .all()
    }

    query = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.deleted_at.is_(None),
            ~P4xTransaction.iban.in_(partner_ibans) if partner_ibans else sa_true(),
        )
        .order_by(P4xTransaction.booking.desc())
    )
    total = query.count()
    items = query.limit(limit).all() if limit else query.all()
    return items, total


def get_warnings_category(
    db: Session,
    limit: int | None = None,
) -> tuple[list[P4xTransaction], int]:
    tx_with_directs = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }

    all_tx = db.query(P4xTransaction).filter(P4xTransaction.deleted_at.is_(None)).all()

    warnings: list[P4xTransaction] = []
    for tx in all_tx:
        if tx.id in tx_with_directs:
            continue
        filter_count = (
            db.query(P4xCategoryFilterHit)
            .filter(P4xCategoryFilterHit.p4x_transaction_id == tx.id)
            .count()
        )
        if filter_count != 1:
            warnings.append(tx)

    total = len(warnings)
    warnings.sort(key=lambda t: t.booking or date.min, reverse=True)
    if limit:
        warnings = warnings[:limit]
    return warnings, total


# ---------------------------------------------------------------------------
# Category filter engine
# ---------------------------------------------------------------------------


def apply_single_filter(db: Session, category_filter: P4xCategoryFilter) -> None:
    tx_with_directs = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }
    _apply_filter_with_cache(db, category_filter, tx_with_directs)
    db.commit()


def _apply_category_filter_criteria(
    query: RowReturningQuery[tuple[int]],
    category_filter: P4xCategoryFilter,
    tx_with_directs: set[int],
) -> RowReturningQuery[tuple[int]]:
    if tx_with_directs:
        query = query.filter(~P4xTransaction.id.in_(tx_with_directs))

    if category_filter.p4x_account_id:
        query = query.filter(
            P4xTransaction.p4x_account_id == category_filter.p4x_account_id
        )

    if category_filter.iban and len(category_filter.iban):
        query = query.filter(P4xTransaction.iban == category_filter.iban)

    if category_filter.min_amount is not None:
        query = query.filter(P4xTransaction.amount >= category_filter.min_amount)

    if category_filter.max_amount is not None:
        query = query.filter(P4xTransaction.amount <= category_filter.max_amount)

    if category_filter.subject is not None:
        query = _apply_subject_filter(query, category_filter)

    return query


def _apply_subject_filter(
    query: RowReturningQuery[tuple[int]],
    category_filter: P4xCategoryFilter,
) -> RowReturningQuery[tuple[int]]:
    if category_filter.subject_mode == "equals":
        return query.filter(P4xTransaction.subject.ilike(category_filter.subject))
    if category_filter.subject_mode == "contains":
        return query.filter(
            P4xTransaction.subject.ilike(f"%{category_filter.subject}%")
        )
    if category_filter.subject_mode == "starts":
        return query.filter(P4xTransaction.subject.ilike(f"{category_filter.subject}%"))
    return query


def _apply_filter_with_cache(
    db: Session,
    category_filter: P4xCategoryFilter,
    tx_with_directs: set[int],
) -> None:
    db.query(P4xCategoryFilterHit).filter(
        P4xCategoryFilterHit.p4x_category_filter_id == category_filter.id,
    ).delete()

    query = db.query(P4xTransaction.id).filter(
        P4xTransaction.deleted_at.is_(None),
    )
    query = _apply_category_filter_criteria(query, category_filter, tx_with_directs)

    now = datetime.now(UTC)
    for (tx_id,) in query.all():
        db.add(
            P4xCategoryFilterHit(
                p4x_transaction_id=tx_id,
                p4x_category_filter_id=category_filter.id,
                created_at=now,
                updated_at=now,
            )
        )


def apply_all_category_filters(
    db: Session,
    *,
    truncate_first: bool = False,
) -> None:
    if truncate_first:
        db.query(P4xCategoryFilterHit).delete()
        db.commit()

    tx_with_directs = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }

    for f in db.query(P4xCategoryFilter).all():
        _apply_filter_with_cache(db, f, tx_with_directs)

    db.commit()


# ---------------------------------------------------------------------------
# Account categories (which categories are used in an account)
# ---------------------------------------------------------------------------


def get_account_categories(db: Session, account: P4xAccount) -> list[P4xCategory]:
    tx_ids = [
        r[0]
        for r in db.query(P4xTransaction.id)
        .filter(
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
        )
        .all()
    ]
    if not tx_ids:
        return []

    direct_cat_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_category_id)
        .filter(
            P4xCategoryDirect.p4x_transaction_id.in_(tx_ids),
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }

    filter_ids = {
        r[0]
        for r in db.query(P4xCategoryFilterHit.p4x_category_filter_id)
        .filter(
            P4xCategoryFilterHit.p4x_transaction_id.in_(tx_ids),
        )
        .distinct()
        .all()
    }
    filter_cat_ids = (
        {
            r[0]
            for r in db.query(P4xCategoryFilter.p4x_category_id)
            .filter(
                P4xCategoryFilter.id.in_(filter_ids),
            )
            .distinct()
            .all()
        }
        if filter_ids
        else set()
    )

    all_cat_ids = direct_cat_ids | filter_cat_ids
    if not all_cat_ids:
        return []

    return db.query(P4xCategory).filter(P4xCategory.id.in_(all_cat_ids)).all()


# ---------------------------------------------------------------------------
# Category CRUD
# ---------------------------------------------------------------------------


def get_category_usage(db: Session, category: P4xCategory) -> dict[str, int]:
    filter_count = (
        db.query(P4xCategoryFilter)
        .filter(
            P4xCategoryFilter.p4x_category_id == category.id,
        )
        .count()
    )
    direct_count = (
        db.query(P4xCategoryDirect)
        .filter(
            P4xCategoryDirect.p4x_category_id == category.id,
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .count()
    )
    return {"filter": filter_count, "direct": direct_count}


def delete_category(db: Session, category: P4xCategory) -> str | None:
    """Returns an error message string if deletion is blocked, else None."""
    if category.protected:
        return "Geschützte Kategorien können nicht gelöscht werden."
    usage = get_category_usage(db, category)
    if usage["filter"] > 0 or usage["direct"] > 0:
        return "Kategorie ist in Verwendung und kann nicht gelöscht werden."
    db.delete(category)
    db.commit()
    return None


# ---------------------------------------------------------------------------
# Category filter CRUD helpers
# ---------------------------------------------------------------------------


def get_filter_hit_count(db: Session, category_filter: P4xCategoryFilter) -> int:
    tx_with_directs = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }

    query = db.query(P4xCategoryFilterHit).filter(
        P4xCategoryFilterHit.p4x_category_filter_id == category_filter.id,
    )
    if tx_with_directs:
        query = query.filter(
            ~P4xCategoryFilterHit.p4x_transaction_id.in_(tx_with_directs),
        )
    return query.count()


def get_filter_hits(
    db: Session,
    category_filter: P4xCategoryFilter,
) -> list[P4xTransaction]:
    tx_with_directs = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }

    hit_tx_ids = [
        r[0]
        for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
        .filter(
            P4xCategoryFilterHit.p4x_category_filter_id == category_filter.id,
        )
        .all()
    ]

    valid_ids = [tid for tid in hit_tx_ids if tid not in tx_with_directs]
    if not valid_ids:
        return []

    return (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.id.in_(valid_ids),
            P4xTransaction.deleted_at.is_(None),
        )
        .all()
    )


def delete_category_filter(db: Session, category_filter: P4xCategoryFilter) -> None:
    db.query(P4xCategoryFilterHit).filter(
        P4xCategoryFilterHit.p4x_category_filter_id == category_filter.id,
    ).delete()
    db.delete(category_filter)
    db.commit()


def filter_to_direct(db: Session, category_filter: P4xCategoryFilter) -> str | None:
    """Convert filter hits to direct assignments. Returns error message or None."""
    _, partner_count = get_warnings_partner(db)
    _, category_count = get_warnings_category(db)
    if partner_count + category_count > 0:
        return "Es gibt noch offene Warnungen. Bitte zuerst alle Warnungen beheben."

    hits = get_filter_hits(db, category_filter)
    for tx in hits:
        set_category_direct(
            db,
            tx,
            [
                {
                    "p4x_category_id": category_filter.p4x_category_id,
                    "amount": tx.amount,
                },
            ],
        )

    db.query(P4xCategoryFilterHit).filter(
        P4xCategoryFilterHit.p4x_category_filter_id == category_filter.id,
    ).delete()
    db.commit()
    return None


# ---------------------------------------------------------------------------
# Category direct assignments
# ---------------------------------------------------------------------------


def set_category_direct(
    db: Session,
    transaction: P4xTransaction,
    slots: list[dict[str, object]],
) -> str | None:
    """Set direct category assignments. Returns error message or None."""
    valid_slots = [s for s in slots if s.get("p4x_category_id") and s.get("amount")]

    if valid_slots:
        total = sum(float(str(s["amount"])) for s in valid_slots)
        if round(float(transaction.amount) - total, 2) != 0.0:
            return "Summe der Beträge stimmt nicht mit dem Transaktionsbetrag überein."

    now = datetime.now(UTC)

    db.query(P4xCategoryDirect).filter(
        P4xCategoryDirect.p4x_transaction_id == transaction.id,
        P4xCategoryDirect.deleted_at.is_(None),
    ).update({"deleted_at": now})

    for slot in valid_slots:
        db.add(
            P4xCategoryDirect(
                p4x_transaction_id=transaction.id,
                p4x_category_id=slot["p4x_category_id"],
                amount=float(str(slot["amount"])),
            )
        )

    db.commit()
    return None


def unset_category_direct(db: Session, transaction: P4xTransaction) -> None:
    now = datetime.now(UTC)
    db.query(P4xCategoryDirect).filter(
        P4xCategoryDirect.p4x_transaction_id == transaction.id,
        P4xCategoryDirect.deleted_at.is_(None),
    ).update({"deleted_at": now})
    db.commit()
    apply_all_category_filters(db)


# ---------------------------------------------------------------------------
# Partner search
# ---------------------------------------------------------------------------


def search_partners(db: Session, term: str) -> list[dict[str, Any]]:
    from app.models.contact import Contact
    from app.models.member import Member
    from app.models.p4x_specialcontact import P4xSpecialcontact

    if len(term) < 3:
        return []

    pattern = f"%{term}%"
    results: list[dict[str, Any]] = []

    members = (
        db.query(Member)
        .filter(
            (Member.vorname.ilike(pattern))
            | (Member.nachname.ilike(pattern))
            | (Member.couleurname.ilike(pattern)),
        )
        .all()
    )
    for m in members:
        results.append({"type": "member", "id": m.id, "label": f"Mitglied: {m.cn}"})

    contacts = (
        db.query(Contact)
        .filter(
            Contact.deleted_at.is_(None),
            (Contact.name.ilike(pattern)) | (Contact.couleurname.ilike(pattern)),
        )
        .all()
    )
    for c in contacts:
        results.append({"type": "contact", "id": c.id, "label": f"Kontakt: {c.cn}"})

    specials = (
        db.query(P4xSpecialcontact)
        .filter(
            P4xSpecialcontact.cn.ilike(pattern),
        )
        .all()
    )
    for s in specials:
        results.append({"type": "special", "id": s.id, "label": f"Spezial: {s.cn}"})

    accounts = (
        db.query(P4xAccount)
        .filter(
            P4xAccount.deleted_at.is_(None),
            P4xAccount.label.ilike(pattern),
        )
        .all()
    )
    for a in accounts:
        results.append({"type": "account", "id": a.id, "label": f"Konto: {a.cn}"})

    return results


# ---------------------------------------------------------------------------
# Partner assignment (1:1 from TransactionController::setPartner)
# ---------------------------------------------------------------------------


def find_partner_entity(
    db: Session,
    partner_type: str,
    partner_id: int,
) -> Member | Contact | P4xAccount | P4xSpecialcontact | None:
    from app.models.contact import Contact
    from app.models.member import Member
    from app.models.p4x_specialcontact import P4xSpecialcontact

    if partner_type == "member":
        return db.query(Member).filter(Member.id == partner_id).first()
    if partner_type == "contact":
        return db.query(Contact).filter(Contact.id == partner_id).first()
    if partner_type == "account":
        return db.query(P4xAccount).filter(P4xAccount.id == partner_id).first()
    if partner_type == "special":
        return (
            db.query(P4xSpecialcontact)
            .filter(
                P4xSpecialcontact.id == partner_id,
            )
            .first()
        )
    return None


def set_transaction_partner(
    db: Session,
    transaction: P4xTransaction,
    partner_data: dict[str, str | int] | None,
    has_delegating: bool,  # noqa: FBT001
    delegating_data: dict[str, str | int] | None,
) -> None:
    now = datetime.now(UTC)

    if partner_data:
        p_type = str(partner_data["type"])
        p_id = int(partner_data["id"])
        remote = find_partner_entity(db, p_type, p_id)
        if remote:
            partner = (
                db.query(P4xPartner)
                .filter(
                    P4xPartner.iban == transaction.iban,
                )
                .first()
            )
            if not partner:
                partner = P4xPartner(iban=transaction.iban)
                db.add(partner)
            partner.partner_type = p_type
            partner.partner_id = remote.id
            partner.deleted_at = None
            partner.updated_at = now
            db.flush()
    else:
        db.query(P4xPartner).filter(
            P4xPartner.iban == transaction.iban,
            P4xPartner.deleted_at.is_(None),
        ).update({"deleted_at": now})

    if has_delegating and delegating_data:
        d_type = str(delegating_data["type"])
        d_id = int(delegating_data["id"])
        remote = find_partner_entity(db, d_type, d_id)
        if remote:
            transaction.delegating_partner_type = d_type
            transaction.delegating_partner_id = remote.id
    else:
        transaction.delegating_partner_type = None
        transaction.delegating_partner_id = None

    transaction.updated_at = now
    db.commit()


# ---------------------------------------------------------------------------
# Transaction edit (comment + attachment)
# ---------------------------------------------------------------------------


def update_transaction_meta(
    db: Session,
    transaction: P4xTransaction,
    comment: str | None,
    file_bytes: bytes | None,
    delete_attachment: bool,  # noqa: FBT001
) -> None:
    import base64

    transaction.comment = comment

    if transaction.has_attachment and delete_attachment:
        transaction.attachment = None
    elif not transaction.has_attachment and file_bytes:
        transaction.attachment = base64.b64encode(file_bytes).decode()

    transaction.updated_at = datetime.now(UTC)
    db.commit()


# ---------------------------------------------------------------------------
# Fee config
# ---------------------------------------------------------------------------


def get_all_fees(db: Session) -> list[P4xFee]:
    from app.models.p4x_fee import P4xFee

    return db.query(P4xFee).order_by(P4xFee.start).all()


def fee_for_month(db: Session, target_date: date) -> float:
    """Returns the fee applicable for a given month (latest start <= target)."""
    from app.models.p4x_fee import P4xFee

    first_of_month = target_date.replace(day=1)
    result = (
        db.query(P4xFee.fee)
        .filter(P4xFee.start <= first_of_month)
        .order_by(P4xFee.start.desc())
        .first()
    )
    return float(result[0]) if result else 0.0


def create_fee(
    db: Session,
    year: int,
    month: int,
    fee_amount: float,
) -> tuple[Any | None, str | None]:
    """Returns (fee, None) on success or (None, error_message) on failure."""
    from app.models.p4x_fee import P4xFee

    start = date(year, month, 1)

    if start < datetime.now(UTC).date().replace(day=1):
        return None, "Startmonat muss aktueller Monat sein oder in der Zukunft liegen."

    existing = db.query(P4xFee).filter(P4xFee.start == start).first()
    if existing:
        return None, "Startmonat muss eindeutig sein."

    fee = P4xFee(start=start, fee=fee_amount, protected=False)
    db.add(fee)
    db.commit()
    return fee, None


def delete_fee(db: Session, start_str: str) -> str | None:
    """Returns error message or None on success."""
    from app.models.p4x_fee import P4xFee

    start = date.fromisoformat(start_str[:10])
    fee = (
        db.query(P4xFee)
        .filter(
            P4xFee.start == start,
            P4xFee.protected == False,  # noqa: E712
        )
        .first()
    )
    if not fee:
        return "Eintrag nicht gefunden oder geschützt."
    db.delete(fee)
    db.commit()
    return None


# ---------------------------------------------------------------------------
# Fee members
# ---------------------------------------------------------------------------

FEE_MEMBER_FILTER = {
    "org_id": "vbw",
    "state_id": "up",
    "entlassen": False,
    "verstorben": False,
}


def is_fee_member(member: Member) -> bool:
    return all(
        getattr(member, attr) == value for attr, value in FEE_MEMBER_FILTER.items()
    )


def get_fee_members(db: Session) -> list[Member]:
    from app.models.member import Member

    return (
        db.query(Member)
        .filter(
            Member.org_id == "vbw",
            Member.state_id == "up",
            Member.entlassen == False,  # noqa: E712
            Member.verstorben == False,  # noqa: E712
        )
        .all()
    )


def search_fee_members(db: Session, term: str) -> list[dict[str, Any]]:
    from app.models.member import Member

    if len(term) < 3:
        return []

    pattern = f"%{term}%"
    members = (
        db.query(Member)
        .filter(
            Member.org_id == "vbw",
            Member.state_id == "up",
            Member.entlassen == False,  # noqa: E712
            Member.verstorben == False,  # noqa: E712
            (Member.vorname.ilike(pattern))
            | (Member.nachname.ilike(pattern))
            | (Member.couleurname.ilike(pattern)),
        )
        .all()
    )
    return [{"id": m.id, "label": m.cn} for m in members]


def update_fee_member(
    db: Session,
    member: Member,
    data: dict[str, str | float | bool | None],
) -> None:
    init_date_raw = data["p4x_init_date"]
    if isinstance(init_date_raw, str):
        member.p4x_init_date = date.fromisoformat(init_date_raw[:10])
    else:
        member.p4x_init_date = None

    init_balance_raw = data["p4x_init_balance"]
    member.p4x_init_balance = (
        int(init_balance_raw) if init_balance_raw is not None else None
    )

    freed_raw = data["p4x_freed"]
    member.p4x_freed = bool(freed_raw) if freed_raw is not None else None

    comment_raw = data.get("p4x_comment")
    member.p4x_comment = str(comment_raw) if comment_raw is not None else None

    db.commit()


# ---------------------------------------------------------------------------
# Fee balance calculation (1:1 from Member::feeBalance)
# ---------------------------------------------------------------------------


def _count_months(start_date: date, end_date: date) -> int:
    """Exact replication of Member::countMonths() in PHP.

    PHP: $start_date->firstOfMonth()->diff($end_date->lastOfMonth())
    Then: $diff->y * 12 + $diff->m + $diff->d / 30, rounded.
    """
    first = start_date.replace(day=1)
    if end_date.month == 12:
        last = date(end_date.year + 1, 1, 1) - timedelta(days=1)
    else:
        last = date(end_date.year, end_date.month + 1, 1) - timedelta(days=1)

    # Manual relativedelta: years, months, days between first and last
    years = last.year - first.year
    months = last.month - first.month
    days = last.day - first.day

    if days < 0:
        months -= 1
        # Days in the previous month of 'last'
        prev_month = last.replace(day=1) - timedelta(days=1)
        days += prev_month.day

    if months < 0:
        years -= 1
        months += 12

    total = years * 12 + months + days / 30
    return round(total)


def _get_fee_payments_sum(
    db: Session,
    member_id: int,
    from_date: date,
    to_date: date,
    *,
    inclusive_end: bool = False,
) -> float:
    """Get sum of fee payments for a member in a date range.

    Fee payments = byPartner('member', id)
    AND byCategory(FEE_CATEGORY_ID) AND amount > 0
    """
    partner_ibans = [
        r[0]
        for r in db.query(P4xPartner.iban)
        .filter(
            P4xPartner.partner_type == "member",
            P4xPartner.partner_id == member_id,
            P4xPartner.deleted_at.is_(None),
        )
        .all()
    ]

    if not partner_ibans:
        return 0.0

    direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.p4x_category_id == FEE_CATEGORY_ID,
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }

    filter_ids = [
        r[0]
        for r in db.query(P4xCategoryFilter.id)
        .filter(
            P4xCategoryFilter.p4x_category_id == FEE_CATEGORY_ID,
        )
        .all()
    ]
    filter_tx_ids = (
        {
            r[0]
            for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
            .filter(
                P4xCategoryFilterHit.p4x_category_filter_id.in_(filter_ids),
            )
            .all()
        }
        if filter_ids
        else set()
    )

    all_direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }
    fee_cat_tx_ids = direct_tx_ids | (filter_tx_ids - all_direct_tx_ids)

    if not fee_cat_tx_ids:
        return 0.0

    query = db.query(func.sum(P4xTransaction.amount)).filter(
        P4xTransaction.deleted_at.is_(None),
        P4xTransaction.amount > 0,
        P4xTransaction.id.in_(fee_cat_tx_ids),
        P4xTransaction.booking >= from_date,
        (
            P4xTransaction.iban.in_(partner_ibans)
            & P4xTransaction.delegating_partner_type.is_(None)
        )
        | (
            (P4xTransaction.delegating_partner_type == "member")
            & (P4xTransaction.delegating_partner_id == member_id)
        ),
    )

    if inclusive_end:
        query = query.filter(P4xTransaction.booking <= to_date)
    else:
        query = query.filter(P4xTransaction.booking < to_date)

    result = query.scalar()
    return float(result) if result else 0.0


def _get_fee_payments_list(
    db: Session,
    member_id: int,
    from_date: date,
    to_date: date,
) -> list[dict[str, Any]]:
    """Get individual fee payments as list for the progress view."""
    partner_ibans = [
        r[0]
        for r in db.query(P4xPartner.iban)
        .filter(
            P4xPartner.partner_type == "member",
            P4xPartner.partner_id == member_id,
            P4xPartner.deleted_at.is_(None),
        )
        .all()
    ]

    if not partner_ibans:
        return []

    direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.p4x_category_id == FEE_CATEGORY_ID,
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }

    filter_ids = [
        r[0]
        for r in db.query(P4xCategoryFilter.id)
        .filter(
            P4xCategoryFilter.p4x_category_id == FEE_CATEGORY_ID,
        )
        .all()
    ]
    filter_tx_ids = (
        {
            r[0]
            for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
            .filter(
                P4xCategoryFilterHit.p4x_category_filter_id.in_(filter_ids),
            )
            .all()
        }
        if filter_ids
        else set()
    )

    all_direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }
    fee_cat_tx_ids = direct_tx_ids | (filter_tx_ids - all_direct_tx_ids)

    if not fee_cat_tx_ids:
        return []

    txs = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.deleted_at.is_(None),
            P4xTransaction.amount > 0,
            P4xTransaction.id.in_(fee_cat_tx_ids),
            P4xTransaction.booking >= from_date,
            P4xTransaction.booking <= to_date,
            (
                P4xTransaction.iban.in_(partner_ibans)
                & P4xTransaction.delegating_partner_type.is_(None)
            )
            | (
                (P4xTransaction.delegating_partner_type == "member")
                & (P4xTransaction.delegating_partner_id == member_id)
            ),
        )
        .all()
    )

    return [
        {
            "type": "payment",
            "booking": str(tx.booking),
            "amount": float(tx.amount),
        }
        for tx in txs
    ]


def calculate_fee_balance(  # noqa: C901
    db: Session,
    member: Member,
    start_date_str: str | None = None,
    end_date_str: str | None = None,
) -> FeeBalanceResult | None:
    """Exact replication of Member::feeBalance() in PHP."""
    if not is_fee_member(member):
        return None

    if member.p4x_init_date is None and member.philistrierungsdatum is None:
        return None

    init_date = (
        member.p4x_init_date if member.p4x_init_date else member.philistrierungsdatum
    )
    if init_date is None:
        return None
    if isinstance(init_date, str):
        init_date = date.fromisoformat(init_date[:10])
    init_date = init_date.replace(day=1)

    # Determine start_date
    if start_date_str:
        try:
            start_date = date.fromisoformat(start_date_str[:10]).replace(day=1)
        except ValueError:
            start_date = init_date
    else:
        start_date = init_date

    if start_date < init_date:
        start_date = init_date

    # Determine end_date
    if end_date_str:
        try:
            parsed_end = date.fromisoformat(end_date_str[:10])
            if parsed_end.month == 12:
                end_date = date(parsed_end.year + 1, 1, 1) - timedelta(days=1)
            else:
                end_date = date(parsed_end.year, parsed_end.month + 1, 1) - timedelta(
                    days=1
                )
        except ValueError:
            prev_month = datetime.now(UTC).date().replace(day=1) - timedelta(days=1)
            end_date = prev_month
    else:
        prev_month = datetime.now(UTC).date().replace(day=1) - timedelta(days=1)
        end_date = prev_month

    if end_date < start_date:
        if start_date.month == 12:
            end_date = date(start_date.year + 1, 1, 1) - timedelta(days=1)
        else:
            end_date = date(start_date.year, start_date.month + 1, 1) - timedelta(
                days=1
            )

    # Calculate start_balance
    start_balance = float(member.p4x_init_balance or 0)

    if not member.p4x_freed:
        prev_month_date = start_date.replace(day=1) - timedelta(days=1)
        n_months = _count_months(init_date, prev_month_date)
        for i in range(n_months):
            if init_date.month + i % 12 <= 12:
                year = init_date.year + (init_date.month - 1 + i) // 12
                month = (init_date.month - 1 + i) % 12 + 1
            else:
                year = init_date.year + (init_date.month - 1 + i) // 12
                month = (init_date.month - 1 + i) % 12 + 1
            current = date(year, month, 10)
            start_balance -= fee_for_month(db, current)

    start_balance += _get_fee_payments_sum(
        db,
        member.id,
        init_date,
        start_date,
        inclusive_end=False,
    )

    # Build progress
    progress: list[dict[str, str | float]] = []

    if not member.p4x_freed:
        n_months = _count_months(start_date, end_date)
        for i in range(n_months):
            year = start_date.year + (start_date.month - 1 + i) // 12
            month = (start_date.month - 1 + i) % 12 + 1
            current = date(year, month, 10)
            fee_amount = fee_for_month(db, current)
            progress.append(
                {
                    "type": "fee",
                    "booking": str(current),
                    "amount": -fee_amount,
                }
            )

    progress.extend(
        _get_fee_payments_list(db, member.id, start_date, end_date),
    )

    end_balance = start_balance + sum(float(e["amount"]) for e in progress)

    progress.sort(key=lambda e: str(e["booking"]))

    fee_entries = [e for e in progress if e["type"] == "fee"]
    payment_entries = [e for e in progress if e["type"] == "payment"]

    return {
        "start_date": str(start_date),
        "start_balance": start_balance,
        "count": {
            "fees": len(fee_entries),
            "payments": len(payment_entries),
        },
        "sum": {
            "fees": sum(float(e["amount"]) for e in fee_entries),
            "payments": sum(float(e["amount"]) for e in payment_entries),
        },
        "end_date": str(end_date),
        "end_balance": end_balance,
        "progress": progress,
    }


def get_debtors(db: Session) -> list[dict[str, int | str | float]]:
    fee_members = get_fee_members(db)
    debtors: list[dict[str, int | str | float]] = []

    for member in fee_members:
        balance = calculate_fee_balance(db, member)
        if balance and balance["end_balance"] < 0:
            debtors.append(
                {
                    "id": member.id,
                    "cn": member.cn,
                    "balance": balance["end_balance"],
                }
            )

    debtors.sort(key=lambda d: float(d.get("balance", 0)))
    return debtors


# ---------------------------------------------------------------------------
# SumUp balance
# ---------------------------------------------------------------------------


def get_sumup_balance(db: Session) -> dict[str, Any]:
    account = (
        db.query(P4xAccount)
        .filter(
            P4xAccount.id == SUMUP_ACCOUNT_ID,
            P4xAccount.deleted_at.is_(None),
        )
        .first()
    )
    if not account:
        return {
            "in_count": 0,
            "in_sum": 0,
            "out_count": 0,
            "out_sum": 0,
            "latest": None,
        }

    category = (
        db.query(P4xCategory)
        .filter(
            P4xCategory.name == SUMUP_CATEGORY_NAME,
        )
        .first()
    )
    if not category:
        return {
            "in_count": 0,
            "in_sum": 0,
            "out_count": 0,
            "out_sum": 0,
            "latest": None,
        }

    direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.p4x_category_id == category.id,
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }

    filter_ids = [
        r[0]
        for r in db.query(P4xCategoryFilter.id)
        .filter(
            P4xCategoryFilter.p4x_category_id == category.id,
        )
        .all()
    ]
    filter_tx_ids = (
        {
            r[0]
            for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
            .filter(
                P4xCategoryFilterHit.p4x_category_filter_id.in_(filter_ids),
            )
            .all()
        }
        if filter_ids
        else set()
    )

    all_direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }
    all_tx_ids = direct_tx_ids | (filter_tx_ids - all_direct_tx_ids)

    if not all_tx_ids:
        return {
            "in_count": 0,
            "in_sum": 0,
            "out_count": 0,
            "out_sum": 0,
            "latest": None,
        }

    txs = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.id.in_(all_tx_ids),
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
        )
        .all()
    )

    in_txs = [t for t in txs if t.amount > 0]
    out_txs = [t for t in txs if t.amount < 0]
    latest = max((t.booking for t in txs), default=None) if txs else None

    return {
        "in_count": len(in_txs),
        "in_sum": round(sum(t.amount for t in in_txs), 2),
        "out_count": len(out_txs),
        "out_sum": round(sum(t.amount for t in out_txs), 2),
        "latest": str(latest) if latest else None,
    }


# ---------------------------------------------------------------------------
# Summary XLSX generation
# ---------------------------------------------------------------------------


def generate_summary_xlsx(  # noqa: C901
    db: Session,
    start: date,
    end: date,
) -> tuple[bytes, list[tuple[str, bytes]]]:
    """Generate XLSX summary and extract PDF attachments.

    Returns (xlsx_bytes, [(filename, pdf_bytes), ...]).
    """
    import base64
    import io
    from itertools import count

    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    start = start.replace(day=1)
    if end.month == 12:
        end = date(end.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(end.year, end.month + 1, 1) - timedelta(days=1)

    categories = {c.id: c for c in db.query(P4xCategory).all()}
    attachment_counter = count(1)
    attachments: list[tuple[str, bytes]] = []

    wb = Workbook()

    # --- Sheet 1: Zusammenfassung ---
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Zusammenfassung")
    ws.title = "Zusammenfassung"
    ws.freeze_panes = "A2"

    headers = [
        "Kontoname",
        "IBAN / BIC",
        f"Stand per {start.day}.{start.month}.{start.year}",
        f"Stand per {end.day}.{end.month}.{end.year}",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    ws["C1"].alignment = Alignment(horizontal="right")
    ws["D1"].alignment = Alignment(horizontal="right")

    accounts = [
        a
        for a in db.query(P4xAccount).filter(P4xAccount.deleted_at.is_(None)).all()
        if db.query(P4xTransaction)
        .filter(
            P4xTransaction.p4x_account_id == a.id,
            P4xTransaction.deleted_at.is_(None),
            P4xTransaction.booking >= str(start),
            P4xTransaction.booking <= str(end),
        )
        .first()
    ]

    for a in accounts:
        ws.append(
            [
                a.label,
                f"{a.iban} / {a.bic}",
                get_account_balance(db, a, start),
                get_account_balance(db, a, end),
            ]
        )

    # --- Per-account sheets ---
    col_names = [
        "Buchungsdatum",
        "Gegenstelle",
        "Betrag",
        "Anh.",
        "Kategorie 1",
        "Kategorie 2",
        "Kategorie 3",
        "Betreff",
        "Kommentar",
    ]

    for a in accounts:
        ws_acc = wb.create_sheet(title=(a.label or "")[:31])
        ws_acc.freeze_panes = "H2"
        ws_acc.append(col_names)
        for cell in ws_acc[1]:
            cell.font = Font(bold=True)
        ws_acc["C1"].alignment = Alignment(horizontal="right")
        ws_acc["D1"].alignment = Alignment(horizontal="center")

        txs = (
            db.query(P4xTransaction)
            .filter(
                P4xTransaction.p4x_account_id == a.id,
                P4xTransaction.deleted_at.is_(None),
                P4xTransaction.booking >= str(start),
                P4xTransaction.booking <= str(end),
            )
            .order_by(P4xTransaction.booking)
            .all()
        )

        for tx in txs:
            att_name = ""
            if tx.has_attachment and tx.attachment:
                num = next(attachment_counter)
                att_name = f"Anhang_{num}.pdf"
                try:
                    attachments.append((att_name, base64.b64decode(tx.attachment)))
                except Exception:  # noqa: BLE001
                    logger.warning("Failed to decode attachment for tx %s", tx.id)

            partner_str = _format_partner_for_xlsx(db, tx)
            subject_str = tx.subject + (
                " [ Anlage liegt vor ]" if tx.has_attachment else ""
            )

            directs = [d for d in (tx.category_directs or []) if d.deleted_at is None]
            filter_hits = list(tx.category_filter_hits or [])

            cat_cells: list[str] = ["", "", ""]
            cat_fills: list[PatternFill | None] = [None, None, None]
            cat_fonts: list[Font | None] = [None, None, None]

            if directs:
                for i, d in enumerate(directs[:3]):
                    cat = categories.get(d.p4x_category_id)
                    if cat:
                        label = cat.label
                        if len(directs) > 1:
                            label += f" ({d.amount})"
                        cat_cells[i] = label
                        cat_fills[i] = PatternFill(
                            start_color=cat.background_color.lstrip("#"),
                            end_color=cat.background_color.lstrip("#"),
                            fill_type="solid",
                        )
                        cat_fonts[i] = Font(color=cat.text_color.lstrip("#"))
            elif len(filter_hits) == 1:
                cat = categories.get(filter_hits[0].category_filter.p4x_category_id)
                if cat:
                    cat_cells[0] = cat.label
                    cat_fills[0] = PatternFill(
                        start_color=cat.background_color.lstrip("#"),
                        end_color=cat.background_color.lstrip("#"),
                        fill_type="solid",
                    )
                    cat_fonts[0] = Font(color=cat.text_color.lstrip("#"))

            row_data = [
                str(tx.booking) if tx.booking else "",
                partner_str,
                tx.amount,
                att_name,
                cat_cells[0],
                cat_cells[1],
                cat_cells[2],
                subject_str,
                tx.comment or "",
            ]
            ws_acc.append(row_data)

            row_num = ws_acc.max_row
            amount_cell = ws_acc.cell(row=row_num, column=3)
            amount_cell.number_format = "#,##0.00 €"
            color = "00FF00" if tx.amount >= 0 else "FF0000"
            amount_cell.font = Font(color=color)

            ws_acc.cell(row=row_num, column=4).alignment = Alignment(
                horizontal="center"
            )

            for i in range(3):
                cell = ws_acc.cell(row=row_num, column=5 + i)
                if cat_fills[i]:
                    cell.fill = cat_fills[i]
                if cat_fonts[i]:
                    cell.font = cat_fonts[i]

    # --- MB-Zahlungen sheet ---
    ws_mb = wb.create_sheet(title="MB-Zahlungen")
    ws_mb.freeze_panes = "B2"
    mb_headers = [
        "Name",
        "Voller Name",
        "Start-Datum",
        "Start-Kontostand",
        "Summe angefallener MB",
        "Summe bezahlter MB",
        "End-Datum",
        "End-Kontostand",
        "befreit",
        "Kommentar",
    ]
    ws_mb.append(mb_headers)
    for cell in ws_mb[1]:
        cell.font = Font(bold=True)

    for col_idx in [4, 5, 6, 8]:
        ws_mb.cell(row=1, column=col_idx).alignment = Alignment(horizontal="right")
    ws_mb.cell(row=1, column=9).alignment = Alignment(horizontal="center")

    fee_members = get_fee_members(db)
    for member in fee_members:
        balance = calculate_fee_balance(db, member, str(start), str(end))
        if not balance:
            continue

        balance_start = date.fromisoformat(balance["start_date"])
        balance_end = date.fromisoformat(balance["end_date"])
        if balance_start > end or balance_end < start:
            continue

        freed = bool(member.p4x_freed)
        ws_mb.append(
            [
                member.nachname or "",
                member.cn,
                "" if freed else balance["start_date"],
                0 if freed else balance["start_balance"],
                0 if freed else balance["sum"]["fees"],
                0 if freed else balance["sum"]["payments"],
                "" if freed else balance["end_date"],
                0 if freed else balance["end_balance"],
                "x" if freed else "",
                member.p4x_comment or "",
            ]
        )

        row_num = ws_mb.max_row
        for col_idx in [4, 5, 6, 8]:
            cell = ws_mb.cell(row=row_num, column=col_idx)
            cell.number_format = "#,##0.00 €"
            val = cell.value
            if isinstance(val, (int, float)):
                color = "00FF00" if val >= 0 else "FF0000"
                cell.font = Font(color=color)
        ws_mb.cell(row=row_num, column=9).alignment = Alignment(horizontal="center")

    wb.active = 0

    from openpyxl.utils import get_column_letter

    for ws_auto in wb.worksheets:
        for col_idx in range(1, ws_auto.max_column + 1):
            max_len = 0
            for row in ws_auto.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            letter = get_column_letter(col_idx)
            ws_auto.column_dimensions[letter].width = max_len + 3

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), attachments


def _format_partner_for_xlsx(db: Session, tx: P4xTransaction) -> str:
    type_labels = {
        "member": "Mitglied",
        "contact": "Kontakt",
        "account": "Konto",
        "special": "Spezial",
    }

    if tx.delegating_partner_type and tx.delegating_partner_id:
        entity = find_partner_entity(
            db, tx.delegating_partner_type, tx.delegating_partner_id
        )
        label = type_labels.get(tx.delegating_partner_type, "")
        cn = getattr(entity, "cn", "unknown") if entity else "unknown"
        return f"{label}: {cn}"

    if tx.partner and tx.partner.deleted_at is None:
        entity = find_partner_entity(db, tx.partner.partner_type, tx.partner.partner_id)
        label = type_labels.get(tx.partner.partner_type, "")
        cn = getattr(entity, "cn", "unknown") if entity else "unknown"
        return f"{label}: {cn}"

    return "unknown: unknown"
