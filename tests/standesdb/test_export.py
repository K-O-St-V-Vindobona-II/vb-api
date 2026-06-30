"""Tests für Standesdb Export — Mailing-Liste + Excel."""

from datetime import date
from io import BytesIO

import bcrypt
from openpyxl import load_workbook

from app.models.badge import Badge
from app.models.contact import Contact
from app.models.member import Member
from app.models.member_badge import MemberBadge
from app.models.member_role import MemberRole
from app.models.org import Org
from app.models.role import Role
from app.models.state import State
from app.services.auth_service import create_user_session
from app.services.export_service import (
    format_bool,
    format_fuzzy_date,
)


def _seed(db):
    db.add_all(
        [
            Org(id="vbw", label="VBW", order=1),
            Org(id="vbn", label="VBN", order=2),
            State(id="fu", label="Fux", order=1),
            State(id="bu", label="Bursch", order=2),
            Role(id="standesfuehrer", group="funktion", label="Standesführer", order=1),
            Badge(id=1, name="Fuxenband", group="jubelband", order=1),
            Badge(id=2, name="Ehrenzeichen Gold", group="ehrenzeichen", order=2),
        ]
    )
    db.commit()


def _admin(db):
    hashed = bcrypt.hashpw(b"pw", bcrypt.gensalt()).decode()
    m = Member(
        email="admin@vbw.at",
        auth_password=hashed,
        auth_locked=False,
        vorname="Admin",
        nachname="User",
        org_id="vbw",
        state_id="bu",
        entlassen=False,
        verstorben=False,
        zustellungen="adresse_privat",
    )
    db.add(m)
    db.commit()
    db.add(
        MemberRole(
            member_id=m.id,
            role_id="standesfuehrer",
            startdate=date(2000, 1, 1),
            enddate=None,
        )
    )
    db.commit()
    return m


def _member(db, **overrides):
    defaults = {
        "auth_password": bcrypt.hashpw(b"pw", bcrypt.gensalt()).decode(),
        "auth_locked": True,
        "org_id": "vbw",
        "state_id": "fu",
        "entlassen": False,
        "verstorben": False,
        "zustellungen": "adresse_privat",
    }
    defaults.update(overrides)
    m = Member(**defaults)
    db.add(m)
    db.commit()
    return m


def _contact(db, **overrides):
    defaults = {
        "kontakttyp": "person",
        "name": "Kontakt Test",
        "org_id": "vbw",
        "zustellungen": True,
    }
    defaults.update(overrides)
    c = Contact(**defaults)
    db.add(c)
    db.commit()
    return c


def _headers(_client, db, admin):
    token, _, _ = create_user_session(db, admin)
    return {"Authorization": f"Bearer {token}"}


def _base_payload(**overrides):
    base = {
        "module": "mailing-liste",
        "include_disabled_delivery": False,
        "include_dead": False,
        "include_common_contacts": False,
        "only_without_email": False,
        "vbw_fu": True,
    }
    base.update(overrides)
    return base


# --- Format helpers ---


class TestFormatHelpers:
    def test_format_bool_true(self):
        assert format_bool(val=True) == "ja"
        assert format_bool(1) == "ja"

    def test_format_bool_false(self):
        assert format_bool(val=False) == "nein"
        assert format_bool(0) == "nein"
        assert format_bool(None) == "nein"

    def test_fuzzy_date_accuracy_0(self):
        assert format_fuzzy_date(date(2020, 5, 15), 0) == ""

    def test_fuzzy_date_accuracy_1(self):
        assert format_fuzzy_date(date(2020, 5, 15), 1) == "2020"

    def test_fuzzy_date_accuracy_2(self):
        assert format_fuzzy_date(date(2020, 5, 15), 2) == "Mai 2020"

    def test_fuzzy_date_accuracy_3(self):
        assert format_fuzzy_date(date(2020, 5, 15), 3) == "15. Mai 2020"

    def test_fuzzy_date_january(self):
        assert format_fuzzy_date(date(2020, 1, 3), 2) == "Jänner 2020"

    def test_fuzzy_date_none(self):
        assert format_fuzzy_date(None, 3) == ""

    def test_fuzzy_date_string(self):
        assert format_fuzzy_date("2020-05-15", 3) == "15. Mai 2020"


# --- Auth & Permission ---


class TestExportAuth:
    def test_export_config_requires_auth(self, client, db_session):
        resp = client.get("/api/standesdb/export/config")
        assert resp.status_code == 401

    def test_export_requires_auth(self, client, db_session):
        resp = client.post("/api/standesdb/export", json=_base_payload())
        assert resp.status_code == 401

    def test_export_config_returns_data(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        resp = client.get("/api/standesdb/export/config", headers=headers)
        assert resp.status_code == 200
        data = resp.json()
        assert "modules" in data
        assert "orgs" in data
        assert "states" in data
        assert "flags" in data
        assert len(data["modules"]) == 4
        assert len(data["orgs"]) == 2

    def test_unknown_module_rejected(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="nonexistent"),
            headers=headers,
        )
        assert resp.status_code == 422


# --- Filter-Logik ---


class TestExportFilter:
    def test_dsgvo_excludes_dismissed(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="Entlassen",
            nachname="Person",
            email="e@test.at",
            entlassen=True,
        )
        _member(db_session, vorname="Normal", nachname="Person", email="n@test.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="mailing-liste", vbw_fu=True),
            headers=headers,
        )
        assert resp.status_code == 200
        text = resp.content.decode()
        assert "Normal Person" in text
        assert "Entlassen Person" not in text

    def test_excludes_dead_by_default(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="Tot",
            nachname="Person",
            email="t@test.at",
            verstorben=True,
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=True),
            headers=headers,
        )
        text = resp.content.decode()
        assert "Tot Person" not in text

    def test_includes_dead_when_flagged(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="Tot",
            nachname="Person",
            email="t@test.at",
            verstorben=True,
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                vbw_fu=True, include_dead=True, include_disabled_delivery=True
            ),
            headers=headers,
        )
        text = resp.content.decode()
        assert "Tot Person" in text

    def test_excludes_disabled_delivery_by_default(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="Deaktiv",
            nachname="Person",
            email="d@test.at",
            zustellungen="deaktiviert",
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=True),
            headers=headers,
        )
        text = resp.content.decode()
        assert "Deaktiv Person" not in text

    def test_only_without_email(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="Mit", nachname="Email", email="m@test.at")
        _member(db_session, vorname="Ohne", nachname="Email", email=None)

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                module="excel-liste-komplett", vbw_fu=True, only_without_email=True
            ),
            headers=headers,
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Mitglieder"]
        names = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        assert "Email" in names
        assert names.count("Email") == 1

    def test_org_state_filter(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="VBW",
            nachname="Fux",
            email="vf@test.at",
            org_id="vbw",
            state_id="fu",
        )
        _member(
            db_session,
            vorname="VBN",
            nachname="Bursch",
            email="vb@test.at",
            org_id="vbn",
            state_id="bu",
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=True, vbn_bu=False),
            headers=headers,
        )
        text = resp.content.decode()
        assert "VBW Fux" in text
        assert "VBN Bursch" not in text

    def test_contacts_filter(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _contact(db_session, name="VBW Kontakt", email="vk@test.at", org_id="vbw")
        _contact(db_session, name="VBN Kontakt", email="nk@test.at", org_id="vbn")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=False, vbw_contacts=True),
            headers=headers,
        )
        text = resp.content.decode()
        assert "VBW Kontakt" in text
        assert "VBN Kontakt" not in text

    def test_common_contacts(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _contact(db_session, name="Allgemein", email="a@test.at", org_id=None)

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=False, include_common_contacts=True),
            headers=headers,
        )
        text = resp.content.decode()
        assert "Allgemein" in text


# --- Mailing-Liste ---


class TestMailingList:
    def test_format(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="Max",
            nachname="Muster",
            couleurname="Testikus",
            email="max@test.at",
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=True),
            headers=headers,
        )
        assert resp.status_code == 200
        assert "text/plain" in resp.headers["content-type"]
        assert "mailing-liste_" in resp.headers["content-disposition"]
        text = resp.content.decode()
        assert '"Max Muster v/o Testikus" <max@test.at>' in text

    def test_excludes_entries_without_email(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="Ohne", nachname="Email", email=None)

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=True),
            headers=headers,
        )
        text = resp.content.decode()
        assert "Ohne Email" not in text

    def test_includes_contacts(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _contact(db_session, name="Kontakt Person", email="kp@test.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(vbw_fu=False, vbw_contacts=True),
            headers=headers,
        )
        text = resp.content.decode()
        assert '"Kontakt Person" <kp@test.at>' in text


# --- Excel ---


class TestExcelExport:
    def test_two_sheets(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="Max", nachname="Muster", email="m@t.at")
        _contact(db_session, name="Kontakt X", email="k@t.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                module="excel-liste-komplett", vbw_fu=True, vbw_contacts=True
            ),
            headers=headers,
        )
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.headers["content-type"]
        assert "excel-liste-komplett_" in resp.headers["content-disposition"]

        wb = load_workbook(BytesIO(resp.content))
        assert "Mitglieder" in wb.sheetnames
        assert "Kontakte" in wb.sheetnames

    def test_member_headers(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="A", nachname="B", email="a@t.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="excel-liste-komplett", vbw_fu=True),
            headers=headers,
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Mitglieder"]
        h = [ws.cell(row=1, column=c).value for c in range(1, 41)]
        assert h[0] == "ID"
        assert h[2] == "Vorname"
        assert h[3] == "Nachname"
        assert h[7] == "Verbindung"
        assert h[8] == "Status"
        assert h[37] == "Leibbursch/Mater"
        assert h[39] == "Ehrungen"

    def test_contact_headers(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _contact(db_session, name="K", email="k@t.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                module="excel-liste-komplett", vbw_fu=False, vbw_contacts=True
            ),
            headers=headers,
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Kontakte"]
        h = [ws.cell(row=1, column=c).value for c in range(1, 16)]
        assert h[0] == "ID"
        assert h[3] == "Name"
        assert h[7] == "Addresse Ort"
        assert h[14] == "referenz"

    def test_boolean_formatting(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="X",
            nachname="Y",
            email="x@t.at",
            gruender=True,
            verstorben=True,
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                module="excel-liste-komplett",
                vbw_fu=True,
                include_dead=True,
                include_disabled_delivery=True,
            ),
            headers=headers,
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Mitglieder"]
        row2 = {
            ws.cell(row=1, column=c).value: ws.cell(row=2, column=c).value
            for c in range(1, 41)
            if ws.cell(row=1, column=c).value
        }
        assert row2.get("Gruender") == "ja"
        assert row2.get("Verstorben") == "ja"

    def test_state_label_formatting(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="X", nachname="Y", email="x@t.at", state_id="fu")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="excel-liste-komplett", vbw_fu=True),
            headers=headers,
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Mitglieder"]
        status_col = None
        for c in range(1, 41):
            if ws.cell(row=1, column=c).value == "Status":
                status_col = c
                break
        assert ws.cell(row=2, column=status_col).value == "Fux"

    def test_badges_formatting(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        m = _member(db_session, vorname="X", nachname="Y", email="x@t.at")
        db_session.add(
            MemberBadge(
                member_id=m.id,
                badge_id=1,
                presentationdate=date(2020, 1, 1),
                presentationdate_accuracy=3,
            )
        )
        db_session.add(
            MemberBadge(
                member_id=m.id,
                badge_id=2,
                presentationdate=date(2021, 6, 15),
                presentationdate_accuracy=3,
            )
        )
        db_session.commit()

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="excel-liste-komplett", vbw_fu=True),
            headers=headers,
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Mitglieder"]
        badges_col = None
        for c in range(1, 41):
            if ws.cell(row=1, column=c).value == "Ehrungen":
                badges_col = c
                break
        val = ws.cell(row=2, column=badges_col).value
        assert "Fuxenband" in val
        assert "Ehrenzeichen Gold" in val

    def test_sorted_by_nachname(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="Z", nachname="Zebra", email="z@t.at")
        _member(db_session, vorname="A", nachname="Adler", email="a@t.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="excel-liste-komplett", vbw_fu=True),
            headers=headers,
        )
        wb = load_workbook(BytesIO(resp.content))
        ws = wb["Mitglieder"]
        names = [ws.cell(row=r, column=4).value for r in range(2, ws.max_row + 1)]
        filtered = [n for n in names if n]
        assert filtered == sorted(filtered)


# --- PDF: Mitgliederverzeichnis ---


class TestBookletExport:
    def test_generates_pdf(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(db_session, vorname="Max", nachname="Muster", email="m@t.at")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="mitgliederverzeichnis", vbw_fu=True),
            headers=headers,
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/pdf"
        assert "mitgliederverzeichnis_" in resp.headers["content-disposition"]
        assert resp.content[:4] == b"%PDF"
        assert len(resp.content) > 1000

    def test_empty_export_still_works(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="mitgliederverzeichnis", vbw_fu=False),
            headers=headers,
        )
        assert resp.status_code == 200
        assert resp.content[:4] == b"%PDF"


# --- PDF: Adress-Etiketten ---


class TestLabelsExport:
    def test_generates_pdf(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _member(
            db_session,
            vorname="Max",
            nachname="Muster",
            email="m@t.at",
            adresse_privat_anschrift="Testgasse 1",
            adresse_privat_plz="1010",
            adresse_privat_ort="Wien",
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(module="adress-etiketten-zweckform-3490", vbw_fu=True),
            headers=headers,
        )
        assert resp.status_code == 200
        assert resp.headers["content-type"] == "application/pdf"
        assert "adress-etiketten-zweckform-3490_" in resp.headers["content-disposition"]
        assert resp.content[:4] == b"%PDF"
        assert len(resp.content) > 500

    def test_includes_contacts(self, client, db_session):
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _contact(
            db_session,
            name="Kontakt Test",
            email="k@t.at",
            adresse_anschrift="Kontaktgasse 5",
            adresse_plz="1020",
            adresse_ort="Wien",
        )

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                module="adress-etiketten-zweckform-3490",
                vbw_fu=False,
                vbw_contacts=True,
            ),
            headers=headers,
        )
        assert resp.status_code == 200
        assert resp.content[:4] == b"%PDF"
