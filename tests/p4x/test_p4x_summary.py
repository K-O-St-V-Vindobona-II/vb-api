import io
from datetime import UTC, date, datetime

from app.models.member import Member
from app.models.org import Org
from app.models.p4x_account import P4xAccount
from app.models.p4x_category import P4xCategory
from app.models.p4x_category_direct import P4xCategoryDirect
from app.models.p4x_fee import P4xFee
from app.models.p4x_partner import P4xPartner
from app.models.p4x_transaction import P4xTransaction
from app.models.state import State
from app.services.p4x_service import generate_summary_xlsx


def _now() -> datetime:
    return datetime.now(UTC)


def _seed(db) -> P4xAccount:
    db.add_all(
        [
            Org(id="vbw", label="VBW", order=1),
            State(id="up", label="UP", order=1),
        ]
    )
    db.commit()

    account = P4xAccount(
        iban="AT942011100005301947",
        bic="GIBAATWWXXX",
        label="Girokonto",
        init_date=date(2015, 1, 1),
        init_balance=0,
        created_at=_now(),
        updated_at=_now(),
    )
    cat = P4xCategory(
        name="eingang.mitgliedsbeitrag",
        label="Mitgliedsbeitrag",
        background_color="#336600",
        text_color="#ffffff",
        created_at=_now(),
        updated_at=_now(),
    )
    db.add_all([account, cat])
    db.commit()
    db.refresh(account)
    db.refresh(cat)

    db.add(P4xFee(start=date(2017, 1, 1), fee=10.0, protected=True))
    db.commit()

    tx = P4xTransaction(
        sha256hash="summary_tx_1",
        booking=date(2026, 3, 15),
        valuation=date(2026, 3, 15),
        iban="DE001",
        amount=15.0,
        subject="MB Kopernikus",
        p4x_account_id=account.id,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(tx)
    db.flush()
    db.add(
        P4xCategoryDirect(
            p4x_transaction_id=tx.id,
            p4x_category_id=cat.id,
            amount=15.0,
        )
    )

    tx2 = P4xTransaction(
        sha256hash="summary_tx_2",
        booking=date(2026, 3, 20),
        valuation=date(2026, 3, 20),
        iban="DE002",
        amount=-28.39,
        subject="Telekom",
        p4x_account_id=account.id,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(tx2)
    db.commit()

    db.add(
        P4xPartner(
            iban="DE001",
            partner_type="member",
            partner_id=1,
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.commit()

    member = Member(
        vorname="Michael",
        nachname="Schimpl",
        couleurname="Kopernikus",
        email="test@test.at",
        auth_password="x",
        auth_locked=False,
        org_id="vbw",
        state_id="up",
        p4x_init_date=date(2017, 1, 1),
        p4x_init_balance=36,
    )
    db.add(member)
    db.commit()

    return account


class TestSummaryXlsx:
    def test_generates_valid_xlsx(self, db_session):
        _seed(db_session)
        xlsx_bytes, attachments = generate_summary_xlsx(
            db_session,
            date(2026, 3, 1),
            date(2026, 3, 31),
        )
        assert len(xlsx_bytes) > 0
        assert isinstance(attachments, list)

    def test_xlsx_has_correct_sheets(self, db_session):
        from openpyxl import load_workbook

        _seed(db_session)
        xlsx_bytes, _ = generate_summary_xlsx(
            db_session,
            date(2026, 3, 1),
            date(2026, 3, 31),
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        sheet_names = wb.sheetnames
        assert "Zusammenfassung" in sheet_names
        assert "MB-Zahlungen" in sheet_names
        assert len(sheet_names) >= 3

    def test_zusammenfassung_has_account(self, db_session):
        from openpyxl import load_workbook

        _seed(db_session)
        xlsx_bytes, _ = generate_summary_xlsx(
            db_session,
            date(2026, 3, 1),
            date(2026, 3, 31),
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Zusammenfassung"]
        values = [cell.value for cell in ws[2]]
        assert "Girokonto" in values

    def test_account_sheet_has_transactions(self, db_session):
        from openpyxl import load_workbook

        _seed(db_session)
        xlsx_bytes, _ = generate_summary_xlsx(
            db_session,
            date(2026, 3, 1),
            date(2026, 3, 31),
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        account_sheet = next(
            s for s in wb.sheetnames if s not in ("Zusammenfassung", "MB-Zahlungen")
        )
        ws = wb[account_sheet]
        assert ws.max_row >= 3

    def test_mb_zahlungen_has_member(self, db_session):
        from openpyxl import load_workbook

        _seed(db_session)
        xlsx_bytes, _ = generate_summary_xlsx(
            db_session,
            date(2026, 3, 1),
            date(2026, 3, 31),
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["MB-Zahlungen"]
        all_values = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            all_values.extend(row)
        assert any("Schimpl" in str(v) for v in all_values if v)

    def test_no_transactions_in_range(self, db_session):
        _seed(db_session)
        xlsx_bytes, _ = generate_summary_xlsx(
            db_session,
            date(2020, 1, 1),
            date(2020, 1, 31),
        )
        assert len(xlsx_bytes) > 0

    def test_attachment_extraction(self, db_session):
        import base64

        account = _seed(db_session)

        tx = P4xTransaction(
            sha256hash="with_attachment",
            booking=date(2026, 3, 25),
            valuation=date(2026, 3, 25),
            iban="AT999",
            amount=-10.0,
            subject="Beleg",
            p4x_account_id=account.id,
            attachment=base64.b64encode(b"%PDF-1.4 test").decode(),
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(tx)
        db_session.commit()

        _, attachments = generate_summary_xlsx(
            db_session,
            date(2026, 3, 1),
            date(2026, 3, 31),
        )
        assert len(attachments) == 1
        assert attachments[0][0].startswith("Anhang_")
        assert attachments[0][0].endswith(".pdf")
        assert attachments[0][1] == b"%PDF-1.4 test"


class TestSummaryXlsxEdgeCases:
    def test_xlsx_with_multiple_direct_categories(self, db_session):
        """Transaction with multiple direct category assignments."""
        from openpyxl import load_workbook

        account = _seed(db_session)

        cat2 = P4xCategory(
            name="eingang.spende",
            label="Spende",
            background_color="#0000ff",
            text_color="#ffffff",
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(cat2)
        db_session.commit()
        db_session.refresh(cat2)

        tx = P4xTransaction(
            sha256hash="multi_direct_tx",
            booking=date(2026, 3, 18),
            valuation=date(2026, 3, 18),
            iban="DE003",
            amount=50.0,
            subject="Gemischte Buchung",
            p4x_account_id=account.id,
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(tx)
        db_session.flush()

        cat1 = db_session.query(P4xCategory).filter(P4xCategory.id == 1).first()
        db_session.add_all(
            [
                P4xCategoryDirect(
                    p4x_transaction_id=tx.id,
                    p4x_category_id=cat1.id,
                    amount=30.0,
                ),
                P4xCategoryDirect(
                    p4x_transaction_id=tx.id,
                    p4x_category_id=cat2.id,
                    amount=20.0,
                ),
            ]
        )
        db_session.commit()

        xlsx_bytes, _ = generate_summary_xlsx(
            db_session, date(2026, 3, 1), date(2026, 3, 31)
        )

        wb = load_workbook(io.BytesIO(xlsx_bytes))
        account_sheet = next(
            s for s in wb.sheetnames if s not in ("Zusammenfassung", "MB-Zahlungen")
        )
        ws = wb[account_sheet]
        # Should have transactions with category columns populated
        assert ws.max_row >= 2

    def test_xlsx_with_invalid_attachment(self, db_session):
        """Attachment that fails base64 decoding should not crash."""
        account = _seed(db_session)

        tx = P4xTransaction(
            sha256hash="bad_attachment",
            booking=date(2026, 3, 22),
            valuation=date(2026, 3, 22),
            iban="AT777",
            amount=-5.0,
            subject="Bad attach",
            p4x_account_id=account.id,
            attachment="!!!not-valid-base64!!!",
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(tx)
        db_session.commit()

        xlsx_bytes, attachments = generate_summary_xlsx(
            db_session, date(2026, 3, 1), date(2026, 3, 31)
        )
        # Should still generate XLSX without crashing
        assert len(xlsx_bytes) > 0
        # Bad attachment should NOT be in the list
        assert len(attachments) == 0

    def test_xlsx_with_filter_hit_category(self, db_session):
        """Transaction with only a filter hit category (no direct)."""
        from openpyxl import load_workbook

        from app.models.p4x_category_filter import P4xCategoryFilter
        from app.models.p4x_category_filter_hit import P4xCategoryFilterHit

        account = _seed(db_session)

        cat = db_session.query(P4xCategory).first()
        cf = P4xCategoryFilter(
            name="xlsx_filter_test",
            p4x_account_id=account.id,
            subject_mode="equals",
            subject="FilterOnly",
            p4x_category_id=cat.id,
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(cf)
        db_session.commit()
        db_session.refresh(cf)

        tx = P4xTransaction(
            sha256hash="filter_only_tx",
            booking=date(2026, 3, 19),
            valuation=date(2026, 3, 19),
            iban="AT888",
            amount=42.0,
            subject="FilterOnly",
            p4x_account_id=account.id,
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(tx)
        db_session.flush()

        db_session.add(
            P4xCategoryFilterHit(
                p4x_transaction_id=tx.id,
                p4x_category_filter_id=cf.id,
                created_at=_now(),
                updated_at=_now(),
            )
        )
        db_session.commit()

        xlsx_bytes, _ = generate_summary_xlsx(
            db_session, date(2026, 3, 1), date(2026, 3, 31)
        )
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        account_sheet = next(
            s for s in wb.sheetnames if s not in ("Zusammenfassung", "MB-Zahlungen")
        )
        ws = wb[account_sheet]
        assert ws.max_row >= 2

    def test_xlsx_december_range(self, db_session):
        """Summary for December crosses year boundary correctly."""
        account = _seed(db_session)

        tx = P4xTransaction(
            sha256hash="dec_tx",
            booking=date(2026, 12, 15),
            valuation=date(2026, 12, 15),
            iban="DE001",
            amount=10.0,
            subject="Dezember Buchung",
            p4x_account_id=account.id,
            created_at=_now(),
            updated_at=_now(),
        )
        db_session.add(tx)
        db_session.commit()

        xlsx_bytes, _ = generate_summary_xlsx(
            db_session, date(2026, 12, 1), date(2026, 12, 31)
        )
        assert len(xlsx_bytes) > 0
