"""Tests for the single-member XLSX statement export
(app/services/p4x_summary_service.py::generate_fee_member_xlsx and the
/p4x/fee-members/{id}/export endpoint).
"""

import io
from datetime import UTC, date, datetime

import bcrypt
from openpyxl import load_workbook

from app.api.router_includes.p4x import _build_fee_member_response
from app.models.member import Member
from app.models.member_role import MemberRole
from app.models.org import Org
from app.models.p4x_account import P4xAccount
from app.models.p4x_category import P4xCategory
from app.models.p4x_fee import P4xFee
from app.models.p4x_partner import P4xPartner
from app.models.p4x_transaction import P4xTransaction
from app.models.role import Role
from app.models.state import State
from app.services.auth_service import create_user_session
from app.services.p4x_summary_service import generate_fee_member_xlsx


def _now() -> datetime:
    return datetime.now(UTC)


def _seed_base(db) -> None:
    db.add_all(
        [
            Org(id="vbw", label="VBW", order=1),
            State(id="up", label="UP", order=1),
            Role(id="phil-xxxx", group="philchc", label="Phil-x", order=1),
        ]
    )
    db.commit()

    db.add(
        P4xCategory(
            id=1,  # FEE_CATEGORY_ID (= 1) is a hardcoded app-level assumption
            name="eingang.mitgliedsbeitrag",
            label="Mitgliedsbeitrag",
            background_color="#336600",
            text_color="#ffffff",
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.add(P4xFee(start=date(2017, 1, 1), fee=10.0, protected=True))
    db.add(
        P4xAccount(
            iban="AT942011100005301947",
            bic="GIBAATWWXXX",
            label="Girokonto",
            init_date=date(2015, 1, 1),
            init_balance=0,
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.commit()


def _create_fee_member(db, couleurname: str = "Idefix") -> Member:
    member = Member(
        vorname="Robert",
        nachname="Hammerl",
        couleurname=couleurname,
        email="export-test@test.at",
        auth_password="x",
        auth_locked=False,
        org_id="vbw",
        state_id="up",
        p4x_init_date=date(2017, 1, 1),
        p4x_init_balance=0,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


def _add_payment(db, member: Member, booking: date, amount: float) -> None:
    account = db.query(P4xAccount).first()

    tx = P4xTransaction(
        sha256_hash=f"export_test_{booking.isoformat()}_{amount}",
        booking=booking,
        valuation=booking,
        iban="DE001",
        amount=amount,
        subject="MB Zahlung",
        p4x_account_id=account.id,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(tx)
    db.commit()
    db.add(
        P4xPartner(
            iban="DE001",
            member_id=member.id,
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.commit()


def _create_admin(db) -> Member:
    pw = bcrypt.hashpw(b"testpass", bcrypt.gensalt()).decode()
    member = Member(
        vorname="Test",
        nachname="Admin",
        couleurname="Tester",
        email="p4x-export-admin@test.at",
        auth_password=pw,
        org_id="vbw",
        state_id="up",
        auth_locked=False,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    db.add(
        MemberRole(
            member_id=member.id,
            role_id="phil-xxxx",
            startdate=date(2020, 1, 1),
            enddate=None,
        )
    )
    db.commit()
    return member


def _create_plain_member(db) -> Member:
    pw = bcrypt.hashpw(b"testpass", bcrypt.gensalt()).decode()
    member = Member(
        vorname="Plain",
        nachname="User",
        email="p4x-export-plain@test.at",
        auth_password=pw,
        org_id="vbw",
        state_id="up",
        auth_locked=False,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


def _login(db, member: Member) -> dict:
    token, _, _ = create_user_session(db, member)
    return {"Authorization": f"Bearer {token}"}


class TestGenerateFeeMemberXlsx:
    def test_has_correct_sheets(self, db_session):
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        response = _build_fee_member_response(db_session, member)

        xlsx_bytes = generate_fee_member_xlsx(response)
        wb = load_workbook(io.BytesIO(xlsx_bytes))

        assert wb.sheetnames == ["Zusammenfassung", "Verlauf"]

    def test_zusammenfassung_content(self, db_session):
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        response = _build_fee_member_response(db_session, member)

        xlsx_bytes = generate_fee_member_xlsx(response)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Zusammenfassung"]

        assert ws["B1"].value == member.cn
        assert "v/o" in member.cn
        assert float(ws["B3"].value) == float(response.balance.start_balance)
        assert float(ws["B7"].value) == float(response.balance.end_balance)

    def test_verlauf_has_running_balance_header(self, db_session):
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        _add_payment(db_session, member, date(2017, 2, 14), 15.0)
        response = _build_fee_member_response(db_session, member)

        xlsx_bytes = generate_fee_member_xlsx(response)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Verlauf"]

        expected_headers = ["Datum", "Transaktionsart", "Betrag", "Saldo"]
        assert [c.value for c in ws[1]] == expected_headers
        assert len(response.balance.progress) >= 2

    def test_verlauf_rows_match_response(self, db_session):
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        _add_payment(db_session, member, date(2017, 2, 14), 15.0)
        response = _build_fee_member_response(db_session, member)

        xlsx_bytes = generate_fee_member_xlsx(response)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Verlauf"]

        for i, entry in enumerate(response.balance.progress, 2):
            cell_date = ws.cell(row=i, column=1).value
            assert cell_date.date() == date.fromisoformat(entry.booking)
            expected_label = "Fälligkeit" if entry.type == "fee" else "Zahlung"
            assert ws.cell(row=i, column=2).value == expected_label
            assert float(ws.cell(row=i, column=3).value) == float(entry.amount)
            assert float(ws.cell(row=i, column=4).value) == float(entry.balance)

        # Last row's Saldo matches Endstand on the Zusammenfassung sheet.
        last_row = len(response.balance.progress) + 1
        ws_summary = wb["Zusammenfassung"]
        assert float(ws.cell(row=last_row, column=4).value) == float(
            ws_summary["B7"].value
        )

    def test_no_progress_uses_literal_endstand(self, db_session):
        _seed_base(db_session)
        member = _create_fee_member(db_session, couleurname="Frei")
        member.p4x_freed = True
        db_session.commit()

        response = _build_fee_member_response(db_session, member)
        assert response.balance is not None
        assert response.balance.progress == []

        xlsx_bytes = generate_fee_member_xlsx(response)
        wb = load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Zusammenfassung"]

        assert float(ws["B7"].value) == float(response.balance.end_balance)


class TestExportFeeMemberEndpoint:
    def test_export_succeeds_and_sanitizes_filename(self, db_session, client):
        _seed_base(db_session)
        admin = _create_admin(db_session)
        headers = _login(db_session, admin)
        member = _create_fee_member(db_session)
        assert "/" in member.cn

        resp = client.get(f"/api/p4x/fee-members/{member.id}/export", headers=headers)

        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        disposition = resp.headers["content-disposition"]
        assert "/" not in disposition.split("filename=", 1)[1]

        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Zusammenfassung", "Verlauf"]

    def test_export_404_for_missing_member(self, db_session, client):
        _seed_base(db_session)
        admin = _create_admin(db_session)
        headers = _login(db_session, admin)

        resp = client.get("/api/p4x/fee-members/99999/export", headers=headers)

        assert resp.status_code == 404

    def test_export_404_when_no_balance_data(self, db_session, client):
        """A member without an init date has no calculable balance to export."""
        _seed_base(db_session)
        admin = _create_admin(db_session)
        headers = _login(db_session, admin)
        no_balance_member = Member(
            vorname="No",
            nachname="Fee",
            email="no-fee@test.at",
            auth_password="x",
            org_id="vbw",
            state_id="up",
            auth_locked=False,
        )
        db_session.add(no_balance_member)
        db_session.commit()
        db_session.refresh(no_balance_member)

        resp = client.get(
            f"/api/p4x/fee-members/{no_balance_member.id}/export", headers=headers
        )

        assert resp.status_code == 404

    def test_export_403_without_permission(self, db_session, client):
        _seed_base(db_session)
        plain = _create_plain_member(db_session)
        headers = _login(db_session, plain)
        member = _create_fee_member(db_session)

        resp = client.get(f"/api/p4x/fee-members/{member.id}/export", headers=headers)

        assert resp.status_code == 403
