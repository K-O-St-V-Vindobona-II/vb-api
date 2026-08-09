"""Tests for the self-service fee-member endpoints
(GET /p4x/fee-members/me and /p4x/fee-members/me/export) - the member-facing
counterpart to the admin GET /p4x/fee-members/{id}[/export] endpoints, but
scoped exclusively to the authenticated member (no id parameter at all) and
without the admin-internal p4x_comment field.
"""

import io
from datetime import UTC, date, datetime

from openpyxl import load_workbook

from app.models.member import Member
from app.models.org import Org
from app.models.p4x_account import P4xAccount
from app.models.p4x_category import P4xCategory
from app.models.p4x_category_direct import P4xCategoryDirect
from app.models.p4x_fee import P4xFee
from app.models.p4x_partner import P4xPartner
from app.models.p4x_transaction import P4xTransaction
from app.models.state import State
from app.services.auth_service import create_user_session

FEE_CATEGORY_ID = 1  # matches the id seeded for P4xCategory in _seed_base


def _now() -> datetime:
    return datetime.now(UTC)


def _seed_base(db) -> None:
    db.add_all(
        [
            Org(id="vbw", label="VBW", order=1),
            State(id="up", label="UP", order=1),
        ]
    )
    db.commit()

    db.add(
        P4xCategory(
            id=1,  # FEE_CATEGORY_ID (= 1) is a hardcoded app-level assumption
            name="eingang.mitgliedsbeitrag",
            label="Mitgliedsbeitrag",
            background_color="#336600",
            text_color="#ffffff",
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.add(P4xFee(start=date(2017, 1, 1), fee=10.0, protected=True))
    db.add(
        P4xAccount(
            iban="AT942011100005301947",
            bic="GIBAATWWXXX",
            label="Girokonto",
            init_date=date(2015, 1, 1),
            init_balance=0,
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.commit()


def _create_fee_member(db, couleurname: str = "Idefix") -> Member:
    """A fee-eligible member with NO p4x role/permission granted - the
    default case for self-service access (must work without any admin
    permission)."""
    member = Member(
        vorname="Robert",
        nachname="Hammerl",
        couleurname=couleurname,
        email=f"self-service-{couleurname.lower()}@test.at",
        auth_password="x",
        auth_locked=False,
        org_id="vbw",
        state_id="up",
        p4x_init_date=date(2017, 1, 1),
        p4x_init_balance=0,
        p4x_comment="Zahlt in Raten - admin-interne Notiz",
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


def _create_non_fee_member(db) -> Member:
    """Fails the is_fee_member() gate on 'entlassen' specifically, keeping
    org_id/state_id identical to a real fee member to isolate that one
    condition."""
    member = Member(
        vorname="Ehemaliges",
        nachname="Mitglied",
        email="non-fee-member@test.at",
        auth_password="x",
        auth_locked=False,
        org_id="vbw",
        state_id="up",
        entlassen=True,
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


def _create_fee_eligible_member_without_balance(db) -> Member:
    """Passes is_fee_member() but has no p4x_init_date - the case
    calculate_fee_balance() cannot produce a balance for."""
    member = Member(
        vorname="Ohne",
        nachname="Kontodaten",
        email="no-balance-member@test.at",
        auth_password="x",
        auth_locked=False,
        org_id="vbw",
        state_id="up",
    )
    db.add(member)
    db.commit()
    db.refresh(member)
    return member


def _add_payment(db, member: Member, booking: date, amount: float) -> None:
    """Records a booked, category-tagged fee payment for the member -
    _get_fee_payments_sum() only counts transactions that are BOTH linked to
    the member via P4xPartner AND tagged with the fee category, so both are
    required for the payment to actually show up in balance.sum.payments."""
    account = db.query(P4xAccount).first()

    tx = P4xTransaction(
        sha256_hash=f"self_service_test_{member.id}_{booking.isoformat()}_{amount}",
        booking=booking,
        valuation=booking,
        iban=f"DE00{member.id}",
        amount=amount,
        subject="MB Zahlung",
        p4x_account_id=account.id,
        created_at=_now(),
        updated_at=_now(),
    )
    db.add(tx)
    db.commit()
    db.add(
        P4xPartner(
            iban=f"DE00{member.id}",
            member_id=member.id,
            created_at=_now(),
            updated_at=_now(),
        )
    )
    db.add(
        P4xCategoryDirect(
            p4x_transaction_id=tx.id,
            p4x_category_id=FEE_CATEGORY_ID,
            amount=amount,
        )
    )
    db.commit()


def _login(db, member: Member) -> dict:
    token, _, _ = create_user_session(db, member)
    return {"Authorization": f"Bearer {token}"}


class TestGetOwnFeeMember:
    def test_returns_own_account_data_for_fee_member(self, db_session, client):
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me", headers=headers)

        assert resp.status_code == 200
        body = resp.json()
        assert body["id"] == member.id
        assert body["cn"] == member.cn
        assert body["balance"] is not None

    def test_p4x_comment_key_absent_from_json(self, db_session, client):
        """Regression test: p4x_comment is admin-internal and must be
        genuinely absent from the response body, not just null."""
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        assert member.p4x_comment  # sanity: the member actually has one set
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me", headers=headers)

        assert resp.status_code == 200
        assert "p4x_comment" not in resp.json()

    def test_plain_member_without_any_p4x_permission_can_call_successfully(
        self, db_session, client
    ):
        """Proves this endpoint is NOT gated behind p4xView/p4xAdmin, unlike
        the admin GET /fee-members/{id} endpoint - _create_fee_member grants
        no role at all."""
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me", headers=headers)

        assert resp.status_code != 403
        assert resp.status_code == 200

    def test_404_for_non_fee_member(self, db_session, client):
        _seed_base(db_session)
        member = _create_non_fee_member(db_session)
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me", headers=headers)

        assert resp.status_code == 404
        assert resp.json()["detail"] == "Kein Beitragsmitglied."

    def test_two_members_each_get_their_own_data(self, db_session, client):
        """Isolation test: since neither endpoint takes an id parameter,
        there is no id to tamper with - instead confirm that two different
        logged-in members each and only ever see their own data through the
        same endpoint."""
        _seed_base(db_session)
        member_a = _create_fee_member(db_session, couleurname="Alpha")
        _add_payment(db_session, member_a, date(2020, 1, 15), 10.0)
        member_b = _create_fee_member(db_session, couleurname="Beta")
        _add_payment(db_session, member_b, date(2020, 1, 15), 20.0)

        resp_a = client.get(
            "/api/p4x/fee-members/me", headers=_login(db_session, member_a)
        )
        resp_b = client.get(
            "/api/p4x/fee-members/me", headers=_login(db_session, member_b)
        )

        assert resp_a.status_code == 200
        assert resp_b.status_code == 200
        body_a = resp_a.json()
        body_b = resp_b.json()
        assert body_a["id"] == member_a.id
        assert body_b["id"] == member_b.id
        assert body_a["id"] != body_b["id"]
        assert (
            body_a["balance"]["sum"]["payments"] != body_b["balance"]["sum"]["payments"]
        )


class TestExportOwnFeeMember:
    def test_export_succeeds_and_returns_xlsx(self, db_session, client):
        _seed_base(db_session)
        member = _create_fee_member(db_session)
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me/export", headers=headers)

        assert resp.status_code == 200
        assert resp.headers["content-type"] == (
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        wb = load_workbook(io.BytesIO(resp.content))
        assert wb.sheetnames == ["Zusammenfassung", "Verlauf"]

    def test_export_404_for_non_fee_member(self, db_session, client):
        _seed_base(db_session)
        member = _create_non_fee_member(db_session)
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me/export", headers=headers)

        assert resp.status_code == 404
        assert resp.json()["detail"] == "Kein Beitragsmitglied."

    def test_export_404_when_fee_eligible_member_has_no_balance_data(
        self, db_session, client
    ):
        _seed_base(db_session)
        member = _create_fee_eligible_member_without_balance(db_session)
        headers = _login(db_session, member)

        resp = client.get("/api/p4x/fee-members/me/export", headers=headers)

        assert resp.status_code == 404
        assert resp.json()["detail"] == "Keine Kontodaten vorhanden."
