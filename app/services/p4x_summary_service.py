from __future__ import annotations

import base64
import io
import logging
from datetime import date, timedelta
from decimal import Decimal
from itertools import count
from typing import TYPE_CHECKING

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet
    from sqlalchemy.orm import Session

from app.models.p4x_account import GIROKONTO_ACCOUNT_ID, P4xAccount
from app.models.p4x_category import P4xCategory
from app.models.p4x_category_direct import P4xCategoryDirect
from app.models.p4x_category_filter import P4xCategoryFilter
from app.models.p4x_category_filter_hit import P4xCategoryFilterHit
from app.models.p4x_transaction import P4xTransaction
from app.schemas.p4x import FeeBalanceResponse, FeeMemberResponse, SumUpBalanceResponse
from app.services import (
    p4x_account_service,
    p4x_fee_balance_service,
    p4x_partner_service,
)

logger = logging.getLogger(__name__)

SUMUP_CATEGORY_NAME = "projekt.bude.sumup"


# ---------------------------------------------------------------------------
# SumUp balance
# ---------------------------------------------------------------------------


def get_sumup_balance(db: Session) -> SumUpBalanceResponse:
    account = (
        db.query(P4xAccount)
        .filter(
            P4xAccount.id == GIROKONTO_ACCOUNT_ID,
            P4xAccount.deleted_at.is_(None),
        )
        .first()
    )
    if not account:
        return SumUpBalanceResponse(
            in_count=0,
            in_sum=Decimal(0),
            out_count=0,
            out_sum=Decimal(0),
            latest=None,
        )

    category = (
        db.query(P4xCategory)
        .filter(
            P4xCategory.name == SUMUP_CATEGORY_NAME,
        )
        .first()
    )
    if not category:
        return SumUpBalanceResponse(
            in_count=0,
            in_sum=Decimal(0),
            out_count=0,
            out_sum=Decimal(0),
            latest=None,
        )

    direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.p4x_category_id == category.id,
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .all()
    }

    filter_ids = [
        r[0]
        for r in db.query(P4xCategoryFilter.id)
        .filter(
            P4xCategoryFilter.p4x_category_id == category.id,
        )
        .all()
    ]
    filter_tx_ids = (
        {
            r[0]
            for r in db.query(P4xCategoryFilterHit.p4x_transaction_id)
            .filter(
                P4xCategoryFilterHit.p4x_category_filter_id.in_(filter_ids),
            )
            .all()
        }
        if filter_ids
        else set()
    )

    all_direct_tx_ids = {
        r[0]
        for r in db.query(P4xCategoryDirect.p4x_transaction_id)
        .filter(
            P4xCategoryDirect.deleted_at.is_(None),
        )
        .distinct()
        .all()
    }
    all_tx_ids = direct_tx_ids | (filter_tx_ids - all_direct_tx_ids)

    if not all_tx_ids:
        return SumUpBalanceResponse(
            in_count=0,
            in_sum=Decimal(0),
            out_count=0,
            out_sum=Decimal(0),
            latest=None,
        )

    txs = (
        db.query(P4xTransaction)
        .filter(
            P4xTransaction.id.in_(all_tx_ids),
            P4xTransaction.p4x_account_id == account.id,
            P4xTransaction.deleted_at.is_(None),
        )
        .all()
    )

    in_txs = [t for t in txs if t.amount > 0]
    out_txs = [t for t in txs if t.amount < 0]
    latest = max((t.booking for t in txs), default=None) if txs else None

    return SumUpBalanceResponse(
        in_count=len(in_txs),
        in_sum=round(sum((t.amount for t in in_txs), start=Decimal(0)), 2),
        out_count=len(out_txs),
        out_sum=round(sum((t.amount for t in out_txs), start=Decimal(0)), 2),
        latest=str(latest) if latest else None,
    )


# ---------------------------------------------------------------------------
# Summary XLSX generation
# ---------------------------------------------------------------------------


def generate_summary_xlsx(  # noqa: C901, PLR0912, PLR0915
    db: Session,
    start: date,
    end: date,
) -> tuple[bytes, list[tuple[str, bytes]]]:
    """Generate XLSX summary and extract PDF attachments.

    Returns (xlsx_bytes, [(filename, pdf_bytes), ...]).
    """
    start = start.replace(day=1)
    if end.month == 12:
        end = date(end.year + 1, 1, 1) - timedelta(days=1)
    else:
        end = date(end.year, end.month + 1, 1) - timedelta(days=1)

    categories = {c.id: c for c in db.query(P4xCategory).all()}
    attachment_counter = count(1)
    attachments: list[tuple[str, bytes]] = []

    wb = Workbook()

    # --- Sheet 1: Zusammenfassung ---
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Zusammenfassung")
    ws.title = "Zusammenfassung"
    ws.freeze_panes = "A2"

    headers = [
        "Kontoname",
        "IBAN / BIC",
        f"Stand per {start.day}.{start.month}.{start.year}",
        f"Stand per {end.day}.{end.month}.{end.year}",
    ]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
    ws["C1"].alignment = Alignment(horizontal="right")
    ws["D1"].alignment = Alignment(horizontal="right")

    accounts = [
        a
        for a in db.query(P4xAccount).filter(P4xAccount.deleted_at.is_(None)).all()
        if db.query(P4xTransaction)
        .filter(
            P4xTransaction.p4x_account_id == a.id,
            P4xTransaction.deleted_at.is_(None),
            P4xTransaction.booking >= str(start),
            P4xTransaction.booking <= str(end),
        )
        .first()
    ]

    for a in accounts:
        ws.append(
            [
                a.label,
                f"{a.iban} / {a.bic}",
                p4x_account_service.get_account_balance(db, a, start),
                p4x_account_service.get_account_balance(db, a, end),
            ]
        )

    # --- Per-account sheets ---
    col_names = [
        "Buchungsdatum",
        "Gegenstelle",
        "Betrag",
        "Anh.",
        "Kategorie 1",
        "Kategorie 2",
        "Kategorie 3",
        "Betreff",
        "Kommentar",
    ]

    for a in accounts:
        ws_acc = wb.create_sheet(title=(a.label or "")[:31])
        ws_acc.freeze_panes = "H2"
        ws_acc.append(col_names)
        for cell in ws_acc[1]:
            cell.font = Font(bold=True)
        ws_acc["C1"].alignment = Alignment(horizontal="right")
        ws_acc["D1"].alignment = Alignment(horizontal="center")

        txs = (
            db.query(P4xTransaction)
            .filter(
                P4xTransaction.p4x_account_id == a.id,
                P4xTransaction.deleted_at.is_(None),
                P4xTransaction.booking >= str(start),
                P4xTransaction.booking <= str(end),
            )
            .order_by(P4xTransaction.booking)
            .all()
        )

        for tx in txs:
            att_name = ""
            if tx.has_attachment and tx.attachment:
                num = next(attachment_counter)
                att_name = f"Anhang_{num}.pdf"
                try:
                    attachments.append((att_name, base64.b64decode(tx.attachment)))
                except Exception:  # noqa: BLE001
                    logger.warning("Failed to decode attachment for tx %s", tx.id)

            partner_str = _format_partner_for_xlsx(db, tx)
            subject_str = tx.subject + (
                " [ Anlage liegt vor ]" if tx.has_attachment else ""
            )

            directs = [d for d in (tx.category_directs or []) if d.deleted_at is None]
            filter_hits = list(tx.category_filter_hits or [])

            cat_cells: list[str] = ["", "", ""]
            cat_fills: list[PatternFill | None] = [None, None, None]
            cat_fonts: list[Font | None] = [None, None, None]

            if directs:
                for i, d in enumerate(directs[:3]):
                    cat = categories.get(d.p4x_category_id)
                    if cat:
                        label = cat.label
                        if len(directs) > 1:
                            label += f" ({d.amount})"
                        cat_cells[i] = label
                        cat_fills[i] = PatternFill(
                            start_color=cat.background_color.lstrip("#"),
                            end_color=cat.background_color.lstrip("#"),
                            fill_type="solid",
                        )
                        cat_fonts[i] = Font(color=cat.text_color.lstrip("#"))
            elif len(filter_hits) == 1:
                cat = categories.get(filter_hits[0].category_filter.p4x_category_id)
                if cat:
                    cat_cells[0] = cat.label
                    cat_fills[0] = PatternFill(
                        start_color=cat.background_color.lstrip("#"),
                        end_color=cat.background_color.lstrip("#"),
                        fill_type="solid",
                    )
                    cat_fonts[0] = Font(color=cat.text_color.lstrip("#"))

            row_data = [
                str(tx.booking) if tx.booking else "",
                partner_str,
                tx.amount,
                att_name,
                cat_cells[0],
                cat_cells[1],
                cat_cells[2],
                subject_str,
                tx.comment or "",
            ]
            ws_acc.append(row_data)

            row_num = ws_acc.max_row
            amount_cell = ws_acc.cell(row=row_num, column=3)
            amount_cell.number_format = "#,##0.00 €"
            color = "00FF00" if tx.amount >= 0 else "FF0000"
            amount_cell.font = Font(color=color)

            ws_acc.cell(row=row_num, column=4).alignment = Alignment(
                horizontal="center"
            )

            for i in range(3):
                cell = ws_acc.cell(row=row_num, column=5 + i)
                if cat_fills[i]:
                    cell.fill = cat_fills[i]
                if cat_fonts[i]:
                    cell.font = cat_fonts[i]

    # --- MB-Zahlungen sheet ---
    ws_mb = wb.create_sheet(title="MB-Zahlungen")
    ws_mb.freeze_panes = "B2"
    mb_headers = [
        "Name",
        "Voller Name",
        "Start-Datum",
        "Start-Kontostand",
        "Summe angefallener MB",
        "Summe bezahlter MB",
        "End-Datum",
        "End-Kontostand",
        "befreit",
        "Kommentar",
    ]
    ws_mb.append(mb_headers)
    for cell in ws_mb[1]:
        cell.font = Font(bold=True)

    for col_idx in [4, 5, 6, 8]:
        ws_mb.cell(row=1, column=col_idx).alignment = Alignment(horizontal="right")
    ws_mb.cell(row=1, column=9).alignment = Alignment(horizontal="center")

    fee_members = p4x_fee_balance_service.get_fee_members(db)
    for member in fee_members:
        balance = p4x_fee_balance_service.calculate_fee_balance(
            db, member, str(start), str(end)
        )
        if not balance:
            continue

        balance_start = date.fromisoformat(balance["start_date"])
        balance_end = date.fromisoformat(balance["end_date"])
        if balance_start > end or balance_end < start:
            continue

        freed = bool(member.p4x_freed)
        ws_mb.append(
            [
                member.nachname or "",
                member.cn,
                "" if freed else balance["start_date"],
                0 if freed else balance["start_balance"],
                0 if freed else balance["sum"]["fees"],
                0 if freed else balance["sum"]["payments"],
                "" if freed else balance["end_date"],
                0 if freed else balance["end_balance"],
                "x" if freed else "",
                member.p4x_comment or "",
            ]
        )

        row_num = ws_mb.max_row
        for col_idx in [4, 5, 6, 8]:
            cell = ws_mb.cell(row=row_num, column=col_idx)
            cell.number_format = "#,##0.00 €"
            val = cell.value
            if isinstance(val, (int, float, Decimal)):
                color = "00FF00" if val >= 0 else "FF0000"
                cell.font = Font(color=color)
        ws_mb.cell(row=row_num, column=9).alignment = Alignment(horizontal="center")

    wb.active = 0

    for ws_auto in wb.worksheets:
        for col_idx in range(1, ws_auto.max_column + 1):
            max_len = 0
            for row in ws_auto.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            letter = get_column_letter(col_idx)
            ws_auto.column_dimensions[letter].width = max_len + 3

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue(), attachments


# ---------------------------------------------------------------------------
# Fee Member Statement XLSX generation
# ---------------------------------------------------------------------------

_AMOUNT_FORMAT = "#,##0.00 €"
_DATE_FORMAT = "DD.MM.YYYY"


def _amount_font(value: Decimal) -> Font:
    return Font(color="008000" if value >= 0 else "FF0000")


def _write_fee_member_summary_sheet(
    ws: Worksheet, cn: str, balance: FeeBalanceResponse
) -> None:
    """Write the "Zusammenfassung" sheet."""
    summary_rows: list[tuple[str, str | Decimal | date]] = [
        ("Mitglied", cn),
        ("Initialdatum", date.fromisoformat(balance.start_date)),
        ("Initialstand", balance.start_balance),
        (f"{balance.count.fees} verrechnete Beiträge", balance.sum.fees),
        (f"{balance.count.payments} geleistete Zahlungen", balance.sum.payments),
        ("Enddatum", date.fromisoformat(balance.end_date)),
    ]
    for row_idx, (label, value) in enumerate(summary_rows, 1):
        ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        cell = ws.cell(row=row_idx, column=2, value=value)
        if isinstance(value, Decimal):
            cell.number_format = _AMOUNT_FORMAT
            cell.alignment = Alignment(horizontal="right")
            cell.font = _amount_font(value)
        elif isinstance(value, date):
            cell.number_format = _DATE_FORMAT
            cell.alignment = Alignment(horizontal="right")

    end_row = len(summary_rows) + 1
    ws.cell(row=end_row, column=1, value="Endstand").font = Font(bold=True)
    end_cell = ws.cell(row=end_row, column=2, value=balance.end_balance)
    end_cell.number_format = _AMOUNT_FORMAT
    end_cell.alignment = Alignment(horizontal="right")
    end_cell.font = _amount_font(balance.end_balance)


def _write_fee_member_progress_sheet(
    ws: Worksheet, balance: FeeBalanceResponse
) -> None:
    """Write the "Verlauf" sheet with a running-balance value per row."""
    ws.freeze_panes = "A2"
    headers = ["Datum", "Transaktionsart", "Betrag", "Saldo"]
    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header).font = Font(bold=True)

    for row_idx, entry in enumerate(balance.progress, 2):
        booking_date = date.fromisoformat(entry.booking)
        date_cell = ws.cell(row=row_idx, column=1, value=booking_date)
        date_cell.number_format = _DATE_FORMAT
        date_cell.alignment = Alignment(horizontal="left")

        type_cell = ws.cell(
            row=row_idx,
            column=2,
            value="Fälligkeit" if entry.type == "fee" else "Zahlung",
        )
        type_cell.alignment = Alignment(horizontal="left")

        amount_cell = ws.cell(row=row_idx, column=3, value=entry.amount)
        amount_cell.number_format = _AMOUNT_FORMAT
        amount_cell.alignment = Alignment(horizontal="right")
        amount_cell.font = _amount_font(entry.amount)

        saldo_cell = ws.cell(row=row_idx, column=4, value=entry.balance)
        saldo_cell.number_format = _AMOUNT_FORMAT
        saldo_cell.alignment = Alignment(horizontal="right")
        saldo_cell.font = _amount_font(entry.balance)


def _autosize_columns(wb: Workbook) -> None:
    for ws_auto in wb.worksheets:
        for col_idx in range(1, ws_auto.max_column + 1):
            max_len = 0
            for row in ws_auto.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value is not None:
                        max_len = max(max_len, len(str(cell.value)))
            letter = get_column_letter(col_idx)
            ws_auto.column_dimensions[letter].width = max_len + 3


def generate_fee_member_xlsx(response: FeeMemberResponse) -> bytes:
    """Generate a two-sheet XLSX statement for one fee member.

    Sheet "Zusammenfassung" mirrors the on-screen summary; sheet "Verlauf"
    lists every booking with a running balance ("Saldo") per row.
    """
    balance = response.balance
    if balance is None:
        msg = "Fee member has no balance data to export."
        raise ValueError(msg)

    wb = Workbook()

    ws_summary = wb.active
    if ws_summary is None:
        ws_summary = wb.create_sheet("Zusammenfassung")
    ws_summary.title = "Zusammenfassung"
    _write_fee_member_summary_sheet(ws_summary, response.cn, balance)

    ws_progress = wb.create_sheet("Verlauf")
    _write_fee_member_progress_sheet(ws_progress, balance)

    wb.active = 0
    _autosize_columns(wb)

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _format_partner_for_xlsx(db: Session, tx: P4xTransaction) -> str:
    type_labels = {
        "member": "Mitglied",
        "contact": "Kontakt",
        "account": "Konto",
        "special": "Spezial",
    }

    if tx.delegating_partner_type and tx.delegating_partner_id:
        entity = p4x_partner_service.find_partner_entity(
            db, tx.delegating_partner_type, tx.delegating_partner_id
        )
        label = type_labels.get(tx.delegating_partner_type, "")
        cn = getattr(entity, "cn", "unknown") if entity else "unknown"
        return f"{label}: {cn}"

    if tx.partner and tx.partner.deleted_at is None:
        entity = p4x_partner_service.find_partner_entity(
            db, tx.partner.partner_type, tx.partner.partner_id
        )
        label = type_labels.get(tx.partner.partner_type, "")
        cn = getattr(entity, "cn", "unknown") if entity else "unknown"
        return f"{label}: {cn}"

    return "unknown: unknown"
