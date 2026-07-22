"""Additional tests for export_service.py to increase code coverage.

Covers: label generation, address formatting, image base64 lookup,
badge classification, org dates, Excel internals (parent_cn, state_label),
pagination logic, RuntimeError guards, and contact-only_without_email filter.
"""

from datetime import UTC, date, datetime
from io import BytesIO
from unittest.mock import MagicMock, patch

import bcrypt
import pytest
from openpyxl import load_workbook

from app.models.badge import Badge
from app.models.contact import Contact
from app.models.member import Member
from app.models.member_badge import MemberBadge
from app.models.member_role import MemberRole
from app.models.org import Org
from app.models.role import Role
from app.models.standesdb_image import StandesdbImage
from app.models.state import State
from app.services.auth_service import create_user_session
from app.services.export_service import (
    _build_org_dates,
    _classify_badges,
    _contact_label_entry,
    _get_image_base64,
    _member_label_entry,
    _prepare_contact_data,
    _prepare_member_data,
    _resolve_delivery_address,
    generate_excel_full,
    generate_labels,
)

# ---------------------------------------------------------------------------
# Shared test-data helpers (same pattern as test_export.py)
# ---------------------------------------------------------------------------


def _seed(db: object) -> None:
    """Insert minimal reference data."""
    db.add_all(
        [
            Org(id="vbw", label="VBW", order=1),
            Org(id="vbn", label="VBN", order=2),
            State(id="fu", label="Fux", order=1),
            State(id="bu", label="Bursch", order=2),
            Role(
                id="standesfuehrer",
                group="funktion",
                label="Standesführer",
                order=1,
            ),
            Badge(id=1, name="Fuxenband", group="jubelband", order=1),
            Badge(id=2, name="Ehrenzeichen Gold", group="ehrenzeichen", order=2),
        ]
    )
    db.commit()


def _admin(db: object) -> Member:
    hashed = bcrypt.hashpw(b"pw", bcrypt.gensalt()).decode()
    m = Member(
        email="admin@vbw.at",
        auth_password=hashed,
        auth_locked=False,
        vorname="Admin",
        nachname="User",
        org_id="vbw",
        state_id="bu",
        entlassen=False,
        verstorben=False,
        zustellungen="adresse_privat",
    )
    db.add(m)
    db.commit()
    db.add(
        MemberRole(
            member_id=m.id,
            role_id="standesfuehrer",
            startdate=date(2000, 1, 1),
            enddate=None,
        )
    )
    db.commit()
    return m


def _member(db: object, **overrides: object) -> Member:
    defaults: dict[str, object] = {
        "auth_password": bcrypt.hashpw(b"pw", bcrypt.gensalt()).decode(),
        "auth_locked": True,
        "org_id": "vbw",
        "state_id": "fu",
        "entlassen": False,
        "verstorben": False,
        "zustellungen": "adresse_privat",
    }
    defaults.update(overrides)
    m = Member(**defaults)
    db.add(m)
    db.commit()
    return m


def _contact(db: object, **overrides: object) -> Contact:
    defaults: dict[str, object] = {
        "kontakttyp": "person",
        "name": "Kontakt Test",
        "org_id": "vbw",
        "zustellungen": True,
    }
    defaults.update(overrides)
    c = Contact(**defaults)
    db.add(c)
    db.commit()
    return c


def _headers(_client_obj: object, db: object, admin: Member) -> dict[str, str]:
    token, _, _ = create_user_session(db, admin)
    return {"Authorization": f"Bearer {token}"}


def _base_payload(**overrides: object) -> dict[str, object]:
    base: dict[str, object] = {
        "module": "mailing-liste",
        "include_disabled_delivery": False,
        "include_dead": False,
        "include_common_contacts": False,
        "only_without_email": False,
        "vbw_fu": True,
    }
    base.update(overrides)
    return base


# ---------------------------------------------------------------------------
# _resolve_delivery_address (L465-476, especially L470)
# ---------------------------------------------------------------------------


class TestResolveDeliveryAddress:
    def test_invalid_zustellungen_returns_empty(self, db_session: object) -> None:
        """zustellungen not in (adresse_privat, adresse_beruf) -> empty strings."""
        _seed(db_session)
        m = _member(
            db_session,
            vorname="A",
            nachname="B",
            zustellungen="deaktiviert",
        )
        anschrift, plz_ort, land = _resolve_delivery_address(m)
        assert anschrift == ""
        assert plz_ort == ""
        assert land == ""

    def test_none_zustellungen_returns_empty(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(db_session, vorname="A", nachname="B", zustellungen=None)
        anschrift, plz_ort, land = _resolve_delivery_address(m)
        assert anschrift == ""
        assert plz_ort == ""
        assert land == ""

    def test_adresse_privat_returns_fields(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(
            db_session,
            vorname="A",
            nachname="B",
            zustellungen="adresse_privat",
            adresse_privat_anschrift="Testgasse 1",
            adresse_privat_plz="1010",
            adresse_privat_ort="Wien",
            adresse_privat_land="Österreich",
        )
        anschrift, plz_ort, land = _resolve_delivery_address(m)
        assert anschrift == "Testgasse 1"
        assert plz_ort == "1010 Wien"
        assert land == "Österreich"

    def test_adresse_beruf_returns_fields(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(
            db_session,
            vorname="A",
            nachname="B",
            zustellungen="adresse_beruf",
            adresse_beruf_anschrift="Bürogasse 5",
            adresse_beruf_plz="1020",
            adresse_beruf_ort="Wien",
            adresse_beruf_land="AT",
        )
        anschrift, plz_ort, land = _resolve_delivery_address(m)
        assert anschrift == "Bürogasse 5"
        assert plz_ort == "1020 Wien"
        assert land == "AT"


# ---------------------------------------------------------------------------
# _build_org_dates (L416-438, especially L437)
# ---------------------------------------------------------------------------


class TestBuildOrgDates:
    def test_member_with_dates_returns_tuples(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(
            db_session,
            vorname="X",
            nachname="Y",
            org_id="vbw",
            aufnahmedatum=date(2020, 3, 15),
            aufnahmedatum_accuracy=3,
            branderdatum=date(2021, 6, 1),
            branderdatum_accuracy=2,
        )
        result = _build_org_dates(m)
        assert len(result) == 2
        assert result[0] == ("15. März 2020", "Reception")
        assert result[1] == ("Juni 2021", "Branderung")

    def test_member_no_dates_returns_empty(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(db_session, vorname="X", nachname="Y")
        result = _build_org_dates(m)
        assert result == []

    def test_vbn_member_uses_filiierung_label(self, db_session: object) -> None:
        """VBN members get 'Filiierung' instead of 'Burschung'."""
        _seed(db_session)
        m = _member(
            db_session,
            vorname="V",
            nachname="N",
            org_id="vbn",
            burschungsdatum=date(2019, 5, 10),
            burschungsdatum_accuracy=3,
        )
        result = _build_org_dates(m)
        assert len(result) == 1
        assert result[0][1] == "Filiierung"

    def test_vbn_member_uses_damenstand_label(self, db_session: object) -> None:
        """VBN members get 'Damenstand' instead of 'Philistrierung'."""
        _seed(db_session)
        m = _member(
            db_session,
            vorname="V",
            nachname="N",
            org_id="vbn",
            philistrierungsdatum=date(2022, 12, 1),
            philistrierungsdatum_accuracy=1,
        )
        result = _build_org_dates(m)
        assert len(result) == 1
        assert result[0][1] == "Damenstand"


# ---------------------------------------------------------------------------
# _classify_badges (L441-462, especially L449-461)
# ---------------------------------------------------------------------------


class TestClassifyBadges:
    def test_classifies_jubelband_and_ehrenzeichen(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(db_session, vorname="Badge", nachname="Test")
        db_session.add(
            MemberBadge(
                member_id=m.id,
                badge_id=1,
                presentationdate=date(2020, 1, 1),
                presentationdate_accuracy=3,
            )
        )
        db_session.add(
            MemberBadge(
                member_id=m.id,
                badge_id=2,
                presentationdate=date(2021, 6, 15),
                presentationdate_accuracy=3,
            )
        )
        db_session.commit()

        jubelbaender, ehrenzeichen = _classify_badges(db_session, m.id)
        assert len(jubelbaender) == 1
        assert jubelbaender[0]["name"] == "Fuxenband"
        assert len(ehrenzeichen) == 1
        assert ehrenzeichen[0]["name"] == "Ehrenzeichen Gold"

    def test_empty_badges_returns_empty_lists(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(db_session, vorname="NoBadge", nachname="Test")
        jubelbaender, ehrenzeichen = _classify_badges(db_session, m.id)
        assert jubelbaender == []
        assert ehrenzeichen == []


# ---------------------------------------------------------------------------
# Member label entry and contact label entry
# ---------------------------------------------------------------------------


class TestLabelEntries:
    def test_member_label_entry_with_address(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(
            db_session,
            vorname="Max",
            nachname="Muster",
            couleurname="Maxl",
            zustellungen="adresse_privat",
            adresse_privat_anschrift="Teststr. 1",
            adresse_privat_plz="1010",
            adresse_privat_ort="Wien",
            adresse_privat_land="AT",
        )
        entry = _member_label_entry(m)
        assert entry["name"] == "Max Muster v/o Maxl"
        assert entry["address"] == "Teststr. 1"
        assert entry["plz_ort"] == "1010 Wien"
        assert entry["land"] == "AT"

    def test_member_label_entry_deaktiviert(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(
            db_session,
            vorname="Max",
            nachname="Muster",
            zustellungen="deaktiviert",
        )
        entry = _member_label_entry(m)
        assert entry["address"] == ""
        assert entry["plz_ort"] == ""
        assert entry["land"] == ""

    def test_contact_label_entry(self, db_session: object) -> None:
        _seed(db_session)
        c = _contact(
            db_session,
            name="Firma Test",
            anrede="Herr",
            adresse_anschrift="Kontaktstr. 5",
            adresse_plz="1020",
            adresse_ort="Wien",
            adresse_land="AT",
        )
        entry = _contact_label_entry(c)
        assert entry["anrede"] == "Herr"
        assert entry["name"] == "Firma Test"
        assert entry["address"] == "Kontaktstr. 5"
        assert entry["plz_ort"] == "1020 Wien"
        assert entry["land"] == "AT"

    def test_contact_label_entry_none_fields(self, db_session: object) -> None:
        _seed(db_session)
        c = _contact(
            db_session,
            name="Minimal",
            anrede=None,
            adresse_anschrift=None,
            adresse_plz=None,
            adresse_ort=None,
            adresse_land=None,
        )
        entry = _contact_label_entry(c)
        assert entry["anrede"] == ""
        assert entry["address"] == ""
        assert entry["plz_ort"] == ""
        assert entry["land"] == ""


# ---------------------------------------------------------------------------
# Get image base64
# ---------------------------------------------------------------------------


class TestGetImageBase64:
    def test_no_image_id_returns_none(self) -> None:
        result = _get_image_base64(MagicMock(), None, MagicMock())
        assert result is None

    def test_deleted_image_returns_none(self, db_session: object) -> None:
        member = Member(vorname="Test", nachname="User")
        db_session.add(member)
        db_session.commit()

        img = StandesdbImage(
            owner_member_id=member.id,
            sha256_hash="del_hash",
            deleted_at=datetime(2024, 1, 1, tzinfo=UTC),
        )
        db_session.add(img)
        db_session.commit()
        result = _get_image_base64(db_session, img.id, MagicMock())
        assert result is None

    def test_image_not_found_returns_none(self, db_session: object) -> None:
        result = _get_image_base64(db_session, 99999, MagicMock())
        assert result is None

    def test_cached_image_returns_base64(self, db_session: object) -> None:
        member = Member(vorname="Test", nachname="User")
        db_session.add(member)
        db_session.commit()

        img = StandesdbImage(
            owner_member_id=member.id,
            sha256_hash="cached_hash",
        )
        db_session.add(img)
        db_session.commit()

        storage = MagicMock()
        storage.exists.return_value = True
        storage.download.return_value = b"fake-image-data"

        result = _get_image_base64(db_session, img.id, storage)
        assert result is not None
        # Should be base64 of b"fake-image-data"
        import base64

        expected = base64.b64encode(b"fake-image-data").decode("ascii")
        assert result == expected

    def test_original_not_found_returns_none(self, db_session: object) -> None:
        member = Member(vorname="Test", nachname="User")
        db_session.add(member)
        db_session.commit()

        img = StandesdbImage(
            owner_member_id=member.id,
            sha256_hash="no_orig_hash",
        )
        db_session.add(img)
        db_session.commit()

        storage = MagicMock()
        # cache miss, original miss
        storage.exists.side_effect = [False, False]

        result = _get_image_base64(db_session, img.id, storage)
        assert result is None

    def test_thumbnail_generation_success(self, db_session: object) -> None:
        member = Member(vorname="Test", nachname="User")
        db_session.add(member)
        db_session.commit()

        img = StandesdbImage(
            owner_member_id=member.id,
            sha256_hash="thumb_hash",
        )
        db_session.add(img)
        db_session.commit()

        storage = MagicMock()
        # cache miss, original exists
        storage.exists.side_effect = [False, True]
        storage.download.return_value = b"original-bytes"

        with patch("app.services.export_service.generate_thumbnail") as mock_thumb:
            mock_thumb.return_value = (b"thumb-bytes", "image/jpeg")
            result = _get_image_base64(db_session, img.id, storage)

        assert result is not None
        import base64

        assert result == base64.b64encode(b"thumb-bytes").decode("ascii")
        storage.upload.assert_called_once()

    def test_thumbnail_generation_oserror_returns_none(
        self, db_session: object
    ) -> None:
        member = Member(vorname="Test", nachname="User")
        db_session.add(member)
        db_session.commit()

        img = StandesdbImage(
            owner_member_id=member.id,
            sha256_hash="err_hash",
        )
        db_session.add(img)
        db_session.commit()

        storage = MagicMock()
        storage.exists.side_effect = [False, True]
        storage.download.return_value = b"bad-data"

        with patch("app.services.export_service.generate_thumbnail") as mock_thumb:
            mock_thumb.side_effect = OSError("corrupt image")
            result = _get_image_base64(db_session, img.id, storage)

        assert result is None


# ---------------------------------------------------------------------------
# _prepare_member_data — parent_id lookup (L486-488)
# ---------------------------------------------------------------------------


class TestPrepareMemberData:
    def test_parent_cn_resolved(self, db_session: object) -> None:
        _seed(db_session)
        parent = _member(
            db_session,
            vorname="Vater",
            nachname="Test",
            couleurname="Papa",
            email="parent@test.at",
        )
        child = _member(
            db_session,
            vorname="Kind",
            nachname="Test",
            email="child@test.at",
            parent_id=parent.id,
        )
        storage = MagicMock()
        storage.exists.return_value = False

        result = _prepare_member_data(db_session, child, storage)
        assert result["parent_cn"] == "Vater Test v/o Papa"

    def test_no_parent_empty_string(self, db_session: object) -> None:
        _seed(db_session)
        m = _member(
            db_session,
            vorname="Solo",
            nachname="Test",
            email="solo@test.at",
        )
        storage = MagicMock()
        storage.exists.return_value = False

        result = _prepare_member_data(db_session, m, storage)
        assert result["parent_cn"] == ""


# ---------------------------------------------------------------------------
# _prepare_contact_data — address formatting (L540-542)
# ---------------------------------------------------------------------------


class TestPrepareContactData:
    def test_contact_address_formatting(self, db_session: object) -> None:
        _seed(db_session)
        c = _contact(
            db_session,
            name="Kontakt Addr",
            adresse_plz="1030",
            adresse_ort="Wien",
        )
        storage = MagicMock()
        storage.exists.return_value = False

        result = _prepare_contact_data(db_session, c, storage)
        assert result["addr_plz_ort"] == "1030 Wien"

    def test_contact_address_none_fields(self, db_session: object) -> None:
        _seed(db_session)
        c = _contact(
            db_session,
            name="Kontakt None",
            adresse_plz=None,
            adresse_ort=None,
        )
        storage = MagicMock()
        storage.exists.return_value = False

        result = _prepare_contact_data(db_session, c, storage)
        assert result["addr_plz_ort"] == ""


# ---------------------------------------------------------------------------
# generate_excel_full — state_id=None (L193) and parent_cn (L202-205)
# ---------------------------------------------------------------------------


class TestExcelFullInternals:
    def test_member_with_no_state_shows_empty(self, db_session: object) -> None:
        """state_id=None triggers the early return in get_state_label (L193)."""
        _seed(db_session)
        m = _member(
            db_session,
            vorname="NoState",
            nachname="Test",
            state_id=None,
        )
        excel_bytes = generate_excel_full(db_session, [m], [])
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Mitglieder"]
        # Find the Status column
        status_col = None
        for c in range(1, 41):
            if ws.cell(row=1, column=c).value == "Status":
                status_col = c
                break
        assert status_col is not None
        # openpyxl stores empty strings as None
        assert ws.cell(row=2, column=status_col).value in ("", None)

    def test_member_with_parent_shows_cn(self, db_session: object) -> None:
        """parent_id set triggers get_parent_cn lookup (L202-205)."""
        _seed(db_session)
        parent = _member(
            db_session,
            vorname="Vater",
            nachname="Test",
            couleurname="Papa",
            email="vater@test.at",
        )
        child = _member(
            db_session,
            vorname="Kind",
            nachname="Test",
            email="kind@test.at",
            parent_id=parent.id,
        )
        excel_bytes = generate_excel_full(db_session, [child], [])
        wb = load_workbook(BytesIO(excel_bytes))
        ws = wb["Mitglieder"]
        # Find the Leibbursch/Mater column
        lb_col = None
        for c in range(1, 41):
            if ws.cell(row=1, column=c).value == "Leibbursch/Mater":
                lb_col = c
                break
        assert lb_col is not None
        val = ws.cell(row=2, column=lb_col).value
        assert "Vater Test" in val


# ---------------------------------------------------------------------------
# generate_labels — pagination (L601-607) and RuntimeError (L614-615)
# ---------------------------------------------------------------------------


class TestGenerateLabelsPagination:
    def test_labels_pagination_fills_page(self, db_session: object) -> None:
        """24 entries fill exactly one page (8 rows x 3 cols = 24)."""
        _seed(db_session)
        admin = _admin(db_session)
        _headers(None, db_session, admin)
        # 24 entries = 8 rows x 3 cols = exactly 1 full page
        # The 25th entry forces a second page
        members = []
        for i in range(25):
            m = _member(
                db_session,
                vorname=f"M{i}",
                nachname=f"N{i}",
                email=f"m{i}@t.at",
                adresse_privat_anschrift=f"Str {i}",
                adresse_privat_plz="1010",
                adresse_privat_ort="Wien",
            )
            members.append(m)

        # Directly call generate_labels; it calls WeasyPrint which
        # produces a real PDF.  This tests the pagination logic at L601-607.
        pdf_bytes = generate_labels(db_session, members, [])
        assert pdf_bytes[:4] == b"%PDF"
        assert len(pdf_bytes) > 500

    def test_labels_with_contacts(self, db_session: object) -> None:
        _seed(db_session)
        contacts = []
        for i in range(5):
            c = _contact(
                db_session,
                name=f"Kontakt {i}",
                adresse_anschrift=f"Gasse {i}",
                adresse_plz="1020",
                adresse_ort="Wien",
            )
            contacts.append(c)
        pdf_bytes = generate_labels(db_session, [], contacts)
        assert pdf_bytes[:4] == b"%PDF"


class TestGenerateLabelsRuntimeError:
    def test_labels_pdf_none_raises(self, db_session: object) -> None:
        """PDF generation returning None triggers RuntimeError (L614-615)."""
        _seed(db_session)
        m = _member(db_session, vorname="X", nachname="Y")
        with patch("app.services.export_service.HTML") as mock_html_cls:
            mock_instance = MagicMock()
            mock_instance.write_pdf.return_value = None
            mock_html_cls.return_value = mock_instance
            with pytest.raises(RuntimeError, match="PDF generation"):
                generate_labels(db_session, [m], [])


# ---------------------------------------------------------------------------
# generate_booklet — RuntimeError guard (L644-645)
# ---------------------------------------------------------------------------


class TestGenerateBookletRuntimeError:
    def test_booklet_pdf_none_raises(self, db_session: object) -> None:
        """PDF generation returning None triggers RuntimeError (L644-645)."""
        _seed(db_session)
        admin = _admin(db_session)
        storage = MagicMock()
        storage.exists.return_value = False

        from app.services.export_service import generate_booklet

        with patch("app.services.export_service.HTML") as mock_html_cls:
            mock_instance = MagicMock()
            mock_instance.write_pdf.return_value = None
            mock_html_cls.return_value = mock_instance
            with pytest.raises(RuntimeError, match="PDF generation"):
                generate_booklet(db_session, [], [], admin, storage)


# ---------------------------------------------------------------------------
# filter_contacts — only_without_email (L162) via endpoint
# ---------------------------------------------------------------------------


class TestFilterContactsOnlyWithoutEmail:
    def test_contacts_only_without_email(
        self, client: object, db_session: object
    ) -> None:
        """only_without_email filters contacts without email (L162)."""
        _seed(db_session)
        admin = _admin(db_session)
        headers = _headers(client, db_session, admin)
        _contact(db_session, name="Mit Email K", email="k@test.at", org_id="vbw")
        _contact(db_session, name="Ohne Email K", email=None, org_id="vbw")

        resp = client.post(
            "/api/standesdb/export",
            json=_base_payload(
                module="mailing-liste",
                vbw_fu=False,
                vbw_contacts=True,
                only_without_email=True,
            ),
            headers=headers,
        )
        assert resp.status_code == 200
        # Mailing list: only entries with email appear
        # But since only_without_email=True filters to contacts WITHOUT email,
        # and mailing list only shows those WITH email, result should be empty
        text = resp.content.decode()
        assert "Mit Email K" not in text
        assert "Ohne Email K" not in text
