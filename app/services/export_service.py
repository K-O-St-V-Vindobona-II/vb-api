import base64
from datetime import UTC, date, datetime
from io import BytesIO
from pathlib import Path

from jinja2 import Environment, FileSystemLoader
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import ColumnElement, or_
from sqlalchemy.orm import Session
from weasyprint import HTML

from app.core.storage import (
    S3_PATH_STANDESDB_CACHE,
    S3_PATH_STANDESDB_IMAGES,
    StorageClient,
    generate_thumbnail,
)
from app.models.contact import Contact
from app.models.member import Member
from app.models.member_badge import MemberBadge
from app.models.org import Org
from app.models.standesdb_image import StandesdbImage
from app.models.state import State

TEMPLATES_DIR = Path(__file__).parent.parent / "templates"

_jinja_env = Environment(
    loader=FileSystemLoader(str(TEMPLATES_DIR)),
    autoescape=True,
)

GERMAN_MONTHS = [
    "",
    "Jänner",
    "Februar",
    "März",
    "April",
    "Mai",
    "Juni",
    "Juli",
    "August",
    "September",
    "Oktober",
    "November",
    "Dezember",
]

AVAILABLE_MODULES = [
    {"id": "mailing-liste", "label": "Mailing-Liste"},
    {"id": "excel-liste-komplett", "label": "Excel-Liste (komplett)"},
    {"id": "mitgliederverzeichnis", "label": "Mitgliederverzeichnis"},
    {
        "id": "adress-etiketten-zweckform-3490",
        "label": "Adress-Etiketten (Zweckform 3490)",
    },
]

FLAGS = {
    "include_disabled_delivery": (
        "Einträge mit deaktivierter Zustellung einbeziehen "
        "(Bei verstorbenen Mitgliedern wird die Zustellung "
        "ebenfalls deaktiviert!)"
    ),
    "include_dead": "Verstorbene Mitglieder einbeziehen",
    "include_common_contacts": ("allgemeine Kontakte einbeziehen"),
    "only_without_email": ("Nur Einträge ohne E-Mail einbeziehen"),
}


def format_fuzzy_date(
    date_val: date | str | None,
    accuracy: int,
) -> str:
    if not date_val or not accuracy:
        return ""
    if isinstance(date_val, str):
        parts = date_val.split("-")
        y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
    else:
        y, m, d = date_val.year, date_val.month, date_val.day
    if accuracy == 1:
        return str(y)
    if accuracy == 2:
        return f"{GERMAN_MONTHS[m]} {y}"
    return f"{d}. {GERMAN_MONTHS[m]} {y}"


def format_bool(val: bool | int | None) -> str:  # noqa: FBT001
    return "ja" if val else "nein"


def get_export_config(db: Session) -> dict[str, object]:
    orgs = db.query(Org).order_by(Org.order).all()
    states = db.query(State).order_by(State.order).all()
    return {
        "modules": AVAILABLE_MODULES,
        "orgs": [{"id": o.id, "label": o.label} for o in orgs],
        "states": [{"id": s.id, "label": s.label} for s in states],
        "flags": FLAGS,
    }


def _build_org_state_conditions(
    db: Session,
    filter_data: dict[str, object],
) -> list[ColumnElement[bool]]:
    orgs = db.query(Org).all()
    states = db.query(State).all()
    conditions = []
    for org in orgs:
        for state in states:
            key = f"{org.id}_{state.id}"
            if filter_data.get(key):
                conditions.append(
                    (Member.org_id == org.id) & (Member.state_id == state.id)
                )
    return conditions


def filter_members(
    db: Session,
    filter_data: dict[str, object],
) -> list[Member]:
    conditions = _build_org_state_conditions(db, filter_data)
    if not conditions:
        return []

    query = db.query(Member).filter(or_(*conditions))
    query = query.filter(Member.entlassen == False)  # noqa: E712

    if not filter_data.get("include_disabled_delivery"):
        query = query.filter(Member.zustellungen != "deaktiviert")
    if not filter_data.get("include_dead"):
        query = query.filter(Member.verstorben == False)  # noqa: E712
    if filter_data.get("only_without_email"):
        query = query.filter(Member.email.is_(None))

    return query.order_by(Member.nachname).all()


def filter_contacts(
    db: Session,
    filter_data: dict[str, object],
) -> list[Contact]:
    orgs = db.query(Org).all()

    conditions = []
    for org in orgs:
        key = f"{org.id}_contacts"
        if filter_data.get(key):
            conditions.append(Contact.org_id == org.id)
    if filter_data.get("include_common_contacts"):
        conditions.append(Contact.org_id.is_(None))

    if not conditions:
        return []

    query = db.query(Contact).filter(or_(*conditions))
    query = query.filter(Contact.deleted_at.is_(None))

    if not filter_data.get("include_disabled_delivery"):
        query = query.filter(Contact.zustellungen == True)  # noqa: E712
    if filter_data.get("only_without_email"):
        query = query.filter(Contact.email.is_(None))

    return query.order_by(Contact.name).all()


def generate_mailing_list(
    members: list[Member],
    contacts: list[Contact],
) -> bytes:
    entries = []
    for m in members:
        email = (m.email or "").strip()
        if email:
            entries.append(f'"{m.cn}" <{email}>')
    for c in contacts:
        email = (c.email or "").strip()
        if email:
            entries.append(f'"{c.cn}" <{email}>')
    return ", \n".join(entries).encode("utf-8")


def generate_excel_full(  # noqa: C901
    db: Session,
    members: list[Member],
    contacts: list[Contact],
) -> bytes:
    state_cache: dict[str, str] = {}
    member_cache: dict[int, str] = {}

    def get_state_label(state_id: str | None) -> str:
        if not state_id:
            return ""
        if state_id not in state_cache:
            s = db.get(State, state_id)
            state_cache[state_id] = (s.label or state_id) if s else state_id
        return state_cache[state_id]

    def get_parent_cn(parent_id: int | None) -> str:
        if not parent_id:
            return ""
        if parent_id not in member_cache:
            p = db.get(Member, parent_id)
            member_cache[parent_id] = p.cn if p else ""
        return member_cache[parent_id]

    def get_badges_str(member: Member) -> str:
        # Badge.name is nullable in the DB; skip any badge without one rather
        # than letting str.join() crash on a None entry.
        badge_names = [
            mb.badge.name for mb in member.member_badges if mb.badge and mb.badge.name
        ]
        return ", ".join(badge_names)

    wb = Workbook()

    ws_members = wb.active
    if ws_members is None:
        msg = "Workbook has no active sheet."
        raise RuntimeError(msg)
    ws_members.title = "Mitglieder"

    member_headers = [
        "ID",
        "Vortitel",
        "Vorname",
        "Nachname",
        "Nachname Geburt",
        "Nachtitel",
        "Couleurname",
        "Verbindung",
        "Status",
        "Entlassen",
        "Verstorben",
        "Sterbedatum",
        "Grabadresse",
        "Geburtsdatum",
        "Aufnahmedatum",
        "Branderdatum",
        "Burschungsdatum",
        "Philistrierungsdatum",
        "Email",
        "Url",
        "Rufnummer Mobil",
        "Rufnummer Privat",
        "Rufnummer Beruf",
        "Zustellungen",
        "Adresse Privat Anschrift",
        "Adresse Privat Plz",
        "Adresse Privat Ort",
        "Adresse Privat Land",
        "Adresse Beruf Anschrift",
        "Adresse Beruf Plz",
        "Adresse Beruf Ort",
        "Adresse Beruf Land",
        "Arbeitgeber",
        "Taetigkeit",
        "Mitgliedschaften",
        "Verbandchargen",
        "Anmerkungen",
        "Leibbursch/Mater",
        "Gruender",
        "Ehrungen",
    ]

    for col_idx, header in enumerate(member_headers, 1):
        cell = ws_members.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    ws_members.freeze_panes = "A2"

    for row_idx, m in enumerate(members, 2):
        vals = [
            str(m.id),
            m.vortitel or "",
            m.vorname or "",
            m.nachname or "",
            m.nachname_geburt or "",
            m.nachtitel or "",
            m.couleurname or "",
            m.org_id or "",
            get_state_label(m.state_id),
            format_bool(m.entlassen),
            format_bool(m.verstorben),
            format_fuzzy_date(m.sterbedatum, m.sterbedatum_accuracy or 0),
            m.grabadresse or "",
            format_fuzzy_date(
                m.geburtsdatum,
                m.geburtsdatum_accuracy or 0,
            ),
            format_fuzzy_date(
                m.aufnahmedatum,
                m.aufnahmedatum_accuracy or 0,
            ),
            format_fuzzy_date(
                m.branderdatum,
                m.branderdatum_accuracy or 0,
            ),
            format_fuzzy_date(
                m.burschungsdatum,
                m.burschungsdatum_accuracy or 0,
            ),
            format_fuzzy_date(
                m.philistrierungsdatum,
                m.philistrierungsdatum_accuracy or 0,
            ),
            m.email or "",
            m.url or "",
            m.rufnummer_mobil or "",
            m.rufnummer_privat or "",
            m.rufnummer_beruf or "",
            m.zustellungen or "",
            m.adresse_privat_anschrift or "",
            m.adresse_privat_plz or "",
            m.adresse_privat_ort or "",
            m.adresse_privat_land or "",
            m.adresse_beruf_anschrift or "",
            m.adresse_beruf_plz or "",
            m.adresse_beruf_ort or "",
            m.adresse_beruf_land or "",
            m.arbeitgeber or "",
            m.taetigkeit or "",
            m.mitgliedschaften or "",
            m.verbandchargen or "",
            m.anmerkungen or "",
            get_parent_cn(m.parent_id),
            format_bool(m.gruender),
            get_badges_str(m),
        ]
        for col_idx, val in enumerate(vals, 1):
            ws_members.cell(row=row_idx, column=col_idx, value=val)

    ws_contacts = wb.create_sheet("Kontakte")

    contact_headers = [
        "ID",
        "Kontakttyp",
        "Anrede",
        "Name",
        "Couleurname",
        "Adresse Anschrift",
        "Adresse Plz",
        "Addresse Ort",
        "Addresse Land",
        "Zustellungen",
        "Email",
        "Rufnummer",
        "Datum",
        "Annmerkungen",
        "referenz",
    ]

    for col_idx, header in enumerate(contact_headers, 1):
        cell = ws_contacts.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    ws_contacts.freeze_panes = "A2"

    for row_idx, c in enumerate(contacts, 2):
        vals = [
            str(c.id),
            c.kontakttyp or "",
            c.anrede or "",
            c.name or "",
            c.couleurname or "",
            c.adresse_anschrift or "",
            c.adresse_plz or "",
            c.adresse_ort or "",
            c.adresse_land or "",
            format_bool(c.zustellungen),
            c.email or "",
            c.rufnummer or "",
            format_fuzzy_date(c.datum, c.datum_accuracy or 0),
            c.anmerkungen or "",
            c.org_id or "",
        ]
        for col_idx, val in enumerate(vals, 1):
            ws_contacts.cell(row=row_idx, column=col_idx, value=val)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _get_image_base64(
    db: Session,
    image_id: int | None,
    storage: StorageClient,
) -> str | None:
    if not image_id:
        return None
    img = db.get(StandesdbImage, image_id)
    if not img or img.deleted_at:
        return None

    cache_key = f"{S3_PATH_STANDESDB_CACHE}/{img.sha256_hash}"
    if storage.exists(cache_key):
        data = storage.download(cache_key)
        return base64.b64encode(data).decode("ascii")

    original_key = f"{S3_PATH_STANDESDB_IMAGES}/{img.sha256_hash}"
    if not storage.exists(original_key):
        return None

    original_data = storage.download(original_key)
    try:
        thumb_bytes, content_type = generate_thumbnail(
            original_data,
            400,
        )
    except (OSError, ValueError):
        return None
    storage.upload(cache_key, thumb_bytes, content_type)
    return base64.b64encode(thumb_bytes).decode("ascii")


def _build_org_dates(member: Member) -> list[tuple[str, str]]:
    date_fields = [
        ("aufnahmedatum", "aufnahmedatum_accuracy", "Reception"),
        ("branderdatum", "branderdatum_accuracy", "Branderung"),
        (
            "burschungsdatum",
            "burschungsdatum_accuracy",
            "Filiierung" if member.org_id == "vbn" else "Burschung",
        ),
        (
            "philistrierungsdatum",
            "philistrierungsdatum_accuracy",
            "Damenstand" if member.org_id == "vbn" else "Philistrierung",
        ),
    ]
    result: list[tuple[str, str]] = []
    for field, acc_field, label in date_fields:
        val = getattr(member, field, None)
        acc = getattr(member, acc_field, 0) or 0
        fmt = format_fuzzy_date(val, acc)
        if fmt:
            result.append((fmt, label))
    return result


def _classify_badges(
    db: Session,
    member_id: int,
) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    badges = db.query(MemberBadge).filter(MemberBadge.member_id == member_id).all()
    jubelbaender: list[dict[str, str]] = []
    ehrenzeichen: list[dict[str, str]] = []
    for mb in badges:
        if not mb.badge:
            continue
        entry: dict[str, str] = {
            "name": mb.badge.name or "",
            "date": format_fuzzy_date(
                mb.presentationdate,
                mb.presentationdate_accuracy or 0,
            ),
        }
        if mb.badge.group == "jubelband":
            jubelbaender.append(entry)
        elif mb.badge.group == "ehrenzeichen":
            ehrenzeichen.append(entry)
    return jubelbaender, ehrenzeichen


def _resolve_delivery_address(
    member: Member,
) -> tuple[str, str, str]:
    zust = member.zustellungen or ""
    if zust not in ("adresse_privat", "adresse_beruf"):
        return "", "", ""
    addr_anschrift = (getattr(member, f"{zust}_anschrift", "") or "").strip()
    plz = (getattr(member, f"{zust}_plz", "") or "").strip()
    ort = (getattr(member, f"{zust}_ort", "") or "").strip()
    addr_plz_ort = f"{plz} {ort}".strip()
    addr_land = (getattr(member, f"{zust}_land", "") or "").strip()
    return addr_anschrift, addr_plz_ort, addr_land


def _prepare_member_data(
    db: Session,
    member: Member,
    storage: StorageClient,
) -> dict[str, object]:
    parent_cn = ""
    if member.parent_id:
        parent = db.get(Member, member.parent_id)
        if parent:
            parent_cn = parent.cn_full

    org_dates = _build_org_dates(member)
    jubelbaender, ehrenzeichen = _classify_badges(db, member.id)
    addr_anschrift, addr_plz_ort, addr_land = _resolve_delivery_address(member)

    return {
        "cn_full": member.cn_full,
        "nachname_geburt": member.nachname_geburt,
        "gruender": member.gruender,
        "verstorben": bool(member.verstorben),
        "state_id": member.state_id or "",
        "org_id": member.org_id or "",
        "grabadresse": member.grabadresse,
        "parent_cn": parent_cn,
        "org_dates": org_dates,
        "addr_anschrift": addr_anschrift,
        "addr_plz_ort": addr_plz_ort,
        "addr_land": addr_land,
        "email": member.email,
        "url": member.url,
        "rufnummer_privat": member.rufnummer_privat,
        "rufnummer_beruf": member.rufnummer_beruf,
        "rufnummer_mobil": member.rufnummer_mobil,
        "arbeitgeber": member.arbeitgeber,
        "taetigkeit": member.taetigkeit,
        "mitgliedschaften": member.mitgliedschaften,
        "verbandchargen": member.verbandchargen,
        "anmerkungen": member.anmerkungen,
        "jubelbaender": jubelbaender,
        "ehrenzeichen": ehrenzeichen,
        "geburtsdatum_fmt": format_fuzzy_date(
            member.geburtsdatum,
            member.geburtsdatum_accuracy or 0,
        ),
        "sterbedatum_fmt": format_fuzzy_date(
            member.sterbedatum,
            member.sterbedatum_accuracy or 0,
        ),
        "img_base64": _get_image_base64(
            db,
            member.default_image,
            storage,
        ),
    }


def _prepare_contact_data(
    db: Session,
    contact: Contact,
    storage: StorageClient,
) -> dict[str, object]:
    plz = (contact.adresse_plz or "").strip()
    ort = (contact.adresse_ort or "").strip()
    return {
        "cn": contact.cn,
        "kontakttyp": contact.kontakttyp or "",
        "org_id": contact.org_id,
        "anrede": contact.anrede,
        "email": contact.email,
        "rufnummer": contact.rufnummer,
        "adresse_anschrift": contact.adresse_anschrift,
        "adresse_land": contact.adresse_land,
        "addr_plz_ort": f"{plz} {ort}".strip(),
        "datum_fmt": format_fuzzy_date(contact.datum, contact.datum_accuracy or 0),
        "anmerkungen": contact.anmerkungen,
        "img_base64": _get_image_base64(
            db,
            contact.default_image,
            storage,
        ),
    }


def _member_label_entry(m: Member) -> dict[str, str | int]:
    anschrift, plz_ort, land = _resolve_delivery_address(m)
    return {
        "anrede": getattr(m, "anrede", None) or "",
        "name": m.cn_full,
        "address": anschrift,
        "plz_ort": plz_ort,
        "land": land,
    }


def _contact_label_entry(c: Contact) -> dict[str, str | int]:
    plz = (c.adresse_plz or "").strip()
    ort = (c.adresse_ort or "").strip()
    return {
        "anrede": c.anrede or "",
        "name": c.name or "",
        "address": (c.adresse_anschrift or "").strip(),
        "plz_ort": f"{plz} {ort}".strip(),
        "land": (c.adresse_land or "").strip(),
    }


def generate_labels(
    _db: Session,
    members: list[Member],
    contacts: list[Contact],
) -> bytes:
    entries: list[dict[str, str | int]] = [_member_label_entry(m) for m in members]
    entries.extend(_contact_label_entry(c) for c in contacts)

    pages: list[list[dict[str, str | int]]] = []
    current_page: list[dict[str, str | int]] = []
    row, col = 0, 0
    for entry in entries:
        entry["row"] = row
        entry["col"] = col
        current_page.append(entry)
        col += 1
        if col > 2:
            col = 0
            row += 1
        if row > 7:
            pages.append(current_page)
            current_page = []
            row, col = 0, 0
    if current_page:
        pages.append(current_page)

    template = _jinja_env.get_template("labels.html")
    html_content = template.render(pages=pages)
    pdf_bytes = HTML(string=html_content).write_pdf()
    if pdf_bytes is None:
        msg = "PDF generation returned no data."
        raise RuntimeError(msg)
    return pdf_bytes


def generate_booklet(
    db: Session,
    members: list[Member],
    contacts: list[Contact],
    current_user: Member,
    storage: StorageClient,
) -> bytes:
    member_data = [_prepare_member_data(db, m, storage) for m in members]
    contact_data = [_prepare_contact_data(db, c, storage) for c in contacts]

    has_vbw = any(m.org_id == "vbw" for m in members)
    has_vbn = any(m.org_id == "vbn" for m in members)

    template = _jinja_env.get_template("booklet.html")
    html_content = template.render(
        has_members=len(members) > 0,
        has_contacts=len(contacts) > 0,
        has_vbw=has_vbw,
        has_vbn=has_vbn,
        today=datetime.now(UTC).date().isoformat(),
        created_by=current_user.cn,
        members=member_data,
        contacts=contact_data,
    )
    pdf_bytes = HTML(string=html_content).write_pdf()
    if pdf_bytes is None:
        msg = "PDF generation returned no data."
        raise RuntimeError(msg)
    return pdf_bytes
